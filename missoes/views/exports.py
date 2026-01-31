"""Export and Import functionality - Excel exports, PDF reports, and bulk imports"""

from django.shortcuts import redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.views.decorators.http import require_POST

from ..models import Oficial, Missao, Designacao, Usuario, Unidade


def exportar_excel(request, tipo):
    """Exporta dados para Excel."""

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse as HR

    wb = openpyxl.Workbook()
    ws = wb.active

    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='8B0000')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def style_header(sheet, num_cols):
        for col in range(1, num_cols + 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    if tipo == 'oficiais':
        ws.title = 'Oficiais'
        ws.append(['CPF', 'RG', 'Nome', 'Nome de Guerra', 'Posto', 'Quadro', 'OBM', 'Função', 'Email', 'Telefone', 'Score'])
        style_header(ws, 11)
        for o in Oficial.objects.all():
            ws.append([o.cpf, o.rg, o.nome, o.nome_guerra, o.posto, o.quadro, o.obm, o.funcao, o.email, o.telefone, o.score])

    elif tipo == 'missoes':
        ws.title = 'Missões'
        ws.append(['ID', 'Tipo', 'Nome', 'Descrição', 'Local', 'Data Início', 'Data Fim', 'Status', 'Documento'])
        style_header(ws, 9)
        for m in Missao.objects.all():
            ws.append([m.id, m.tipo, m.nome, m.descricao, m.local,
                      m.data_inicio.strftime('%Y-%m-%d') if m.data_inicio else '',
                      m.data_fim.strftime('%Y-%m-%d') if m.data_fim else '',
                      m.status, m.documento_referencia])

    elif tipo == 'designacoes':
        ws.title = 'Designações'
        ws.append(['ID', 'ID Missão', 'Nome Missão', 'RG Oficial', 'Nome Oficial', 'Função', 'Complexidade', 'Observações'])
        style_header(ws, 8)

        # Verificar se está consultando outro oficial
        oficial_id = request.GET.get('oficial_id')

        if oficial_id:
            # Consultando outro oficial
            try:
                oficial_consulta = Oficial.objects.get(pk=oficial_id)
                if request.user.pode_ver_oficial(oficial_consulta):
                    designacoes = Designacao.objects.select_related('missao', 'oficial').filter(
                        oficial=oficial_consulta
                    )
                else:
                    designacoes = Designacao.objects.none()
            except Oficial.DoesNotExist:
                designacoes = Designacao.objects.none()
        elif request.user.oficial and not request.user.is_admin:
            # Próprio oficial (não admin)
            designacoes = Designacao.objects.select_related('missao', 'oficial').filter(
                oficial=request.user.oficial
            )
        else:
            # Admin vê todos
            designacoes = Designacao.objects.select_related('missao', 'oficial').all()

        for d in designacoes:
            ws.append([d.id, d.missao.id, d.missao.nome, d.oficial.rg, str(d.oficial),
                      d.funcao.funcao, d.complexidade, d.observacoes])

    elif tipo == 'unidades':
        ws.title = 'Unidades'
        ws.append(['ID', 'Nome', 'Sigla', 'Tipo', 'ID Comando Superior'])
        style_header(ws, 5)
        for u in Unidade.objects.all():
            ws.append([u.id, u.nome, u.sigla, u.tipo, u.comando_superior_id or ''])

    elif tipo == 'usuarios':
        ws.title = 'Usuários'
        ws.append(['ID', 'CPF', 'Perfil', 'RG Oficial', 'Nome Oficial', 'Ativo'])
        style_header(ws, 6)
        for u in Usuario.objects.select_related('oficial').all():
            ws.append([u.id, u.cpf, u.role,
                      u.oficial.rg if u.oficial else '',
                      str(u.oficial) if u.oficial else '',
                      'Sim' if u.is_active else 'Não'])

    elif tipo == 'modelo':
        # Criar planilha modelo com todas as abas
        return gerar_modelo_importacao()

    # Ajustar largura das colunas
    for column_cells in ws.columns:
        length = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

    response = HR(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=sigem_{tipo}.xlsx'
    wb.save(response)

    return response


def gerar_modelo_importacao():
    """Gera planilha modelo para importação."""

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse as HR

    wb = openpyxl.Workbook()

    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='8B0000')
    info_font = Font(bold=True, color='8B0000')

    def setup_sheet(sheet, headers, widths, example, info_col, info_data):
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            sheet.column_dimensions[get_column_letter(col)].width = widths[col-1]

        for col, value in enumerate(example, 1):
            sheet.cell(row=2, column=col, value=value)

        if info_data:
            for row, (title, values) in enumerate(info_data.items(), 1):
                sheet.cell(row=row, column=info_col, value=title).font = info_font
                for i, v in enumerate(values, 1):
                    sheet.cell(row=row+i, column=info_col, value=v)

        sheet.freeze_panes = 'A2'

    # Aba Oficiais
    ws = wb.active
    ws.title = 'Oficiais'
    setup_sheet(ws,
        ['CPF*', 'RG*', 'Nome Completo*', 'Nome de Guerra', 'Posto*', 'Quadro*', 'OBM', 'Função', 'Email', 'Telefone'],
        [15, 15, 35, 20, 12, 15, 25, 25, 30, 18],
        ['12345678901', 'RG123456', 'JOÃO DA SILVA', 'SILVA', 'Cap', 'QOC', '1º BBM', 'Cmt Cia', 'joao@email.com', '62999999999'],
        12,
        {'POSTOS:': ['Cel', 'TC', 'Maj', 'Cap', '1º Ten', '2º Ten', 'Asp'],
         'QUADROS:': ['QOC', 'QOA/Adm', 'QOA/Mús', 'QOM/Médico', 'QOM/Dentista']}
    )

    # Aba Missões
    ws2 = wb.create_sheet('Missoes')
    setup_sheet(ws2,
        ['Tipo*', 'Nome*', 'Descrição', 'Local', 'Data Início', 'Data Fim', 'Status*', 'Documento'],
        [18, 35, 40, 25, 15, 15, 18, 20],
        ['OPERACIONAL', 'Operação Exemplo', 'Descrição da missão', 'Goiânia-GO', '2024-01-15', '2024-01-20', 'EM_ANDAMENTO', 'SEI-123'],
        10,
        {'TIPOS:': ['OPERACIONAL', 'ADMINISTRATIVA', 'ENSINO', 'CORREICIONAL', 'COMISSAO', 'ACAO_SOCIAL'],
         'STATUS:': ['PLANEJADA', 'EM_ANDAMENTO', 'CONCLUIDA', 'CANCELADA']}
    )

    # Aba Designações
    ws3 = wb.create_sheet('Designacoes')
    setup_sheet(ws3,
        ['ID Missão*', 'RG Oficial*', 'Função*', 'Complexidade*', 'Observações'],
        [15, 18, 20, 15, 40],
        [1, 'RG123456', 'COMANDANTE', 'ALTA', 'Observação opcional'],
        7,
        {'FUNÇÕES:': ['COMANDANTE', 'SUBCOMANDANTE', 'COORDENADOR', 'PRESIDENTE', 'MEMBRO', 'AUXILIAR', 'INSTRUTOR', 'ENCARREGADO', 'RELATOR', 'ESCRIVAO'],
         'COMPLEXIDADE:': ['BAIXA', 'MEDIA', 'ALTA']}
    )

    # Aba Unidades
    ws4 = wb.create_sheet('Unidades')
    setup_sheet(ws4,
        ['Nome*', 'Sigla', 'Tipo*', 'ID Cmd Superior'],
        [40, 15, 18, 18],
        ['1º Batalhão BM', '1º BBM', 'BBM', ''],
        6,
        {'TIPOS:': ['COMANDO_GERAL', 'DIRETORIA', 'BBM', 'CIBM', 'CBM', 'SECAO']}
    )

    # Aba Usuários
    ws5 = wb.create_sheet('Usuarios')
    setup_sheet(ws5,
        ['CPF*', 'Perfil*', 'RG Oficial Vinculado'],
        [18, 15, 22],
        ['12345678901', 'oficial', 'RG123456'],
        5,
        {'PERFIS:': ['admin', 'gestor', 'comandante', 'oficial'],
         'NOTA:': ['Senha padrão: 123456']}
    )

    response = HR(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=sigem_modelo_importacao.xlsx'
    wb.save(response)

    return response


@login_required
def exportar_pdf(request, tipo):
    """Exporta dados para PDF - Relatório de designações do oficial."""

    from django.http import HttpResponse
    from django.conf import settings
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from io import BytesIO
    from datetime import datetime
    import os

    # Só permite PDF de designações por enquanto
    if tipo != 'designacoes':
        messages.info(request, 'Exportação PDF disponível apenas para designações.')
        return redirect(request.META.get('HTTP_REFERER', 'dashboard'))

    # Verificar se está consultando outro oficial
    oficial_id = request.GET.get('oficial_id')

    if oficial_id:
        # Consultando outro oficial
        oficial = get_object_or_404(Oficial, pk=oficial_id)

        # Verificar permissão
        if not request.user.pode_ver_oficial(oficial):
            messages.error(request, 'Você não tem permissão para gerar relatório deste oficial.')
            return redirect('consultar_oficial')
    else:
        # Próprio oficial
        if not request.user.oficial:
            messages.error(request, 'Usuário não vinculado a um oficial.')
            return redirect('consultar_oficial')
        oficial = request.user.oficial

    # Buscar designações do oficial
    designacoes = Designacao.objects.select_related('missao').filter(
        oficial=oficial
    ).order_by('-missao__status', '-criado_em')

    # Criar PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=1.5*cm,
        bottomMargin=1.5*cm
    )

    elements = []
    styles = getSampleStyleSheet()

    # ============================================================
    # FUNÇÃO AUXILIAR: Redimensionar imagem mantendo proporção
    # ============================================================
    def get_image_with_aspect_ratio(img_path, max_width, max_height, preserve_transparency=False):
        """Retorna Image do ReportLab mantendo proporção."""
        from PIL import Image as PILImage
        from PIL import ExifTags

        pil_img = PILImage.open(img_path)

        # Corrigir orientação EXIF
        try:
            for orientation in ExifTags.TAGS.keys():
                if ExifTags.TAGS[orientation] == 'Orientation':
                    break

            exif = pil_img._getexif()
            if exif is not None:
                orientation_value = exif.get(orientation)

                if orientation_value == 2:
                    pil_img = pil_img.transpose(PILImage.FLIP_LEFT_RIGHT)
                elif orientation_value == 3:
                    pil_img = pil_img.rotate(180)
                elif orientation_value == 4:
                    pil_img = pil_img.rotate(180).transpose(PILImage.FLIP_LEFT_RIGHT)
                elif orientation_value == 5:
                    pil_img = pil_img.rotate(-90, expand=True).transpose(PILImage.FLIP_LEFT_RIGHT)
                elif orientation_value == 6:
                    pil_img = pil_img.rotate(-90, expand=True)
                elif orientation_value == 7:
                    pil_img = pil_img.rotate(90, expand=True).transpose(PILImage.FLIP_LEFT_RIGHT)
                elif orientation_value == 8:
                    pil_img = pil_img.rotate(90, expand=True)
        except (AttributeError, KeyError, IndexError):
            pass

        # Obter dimensões originais
        orig_width, orig_height = pil_img.size

        # Calcular proporção mantendo aspect ratio
        ratio = min(max_width / orig_width, max_height / orig_height)
        new_width = orig_width * ratio
        new_height = orig_height * ratio

        # Salvar em buffer
        img_buffer = BytesIO()
        if preserve_transparency and pil_img.mode in ('RGBA', 'P', 'LA'):
            # Manter transparência - salvar como PNG
            pil_img.save(img_buffer, format='PNG')
        else:
            # Sem transparência - salvar como JPEG
            if pil_img.mode in ('RGBA', 'P', 'LA'):
                pil_img = pil_img.convert('RGB')
            pil_img.save(img_buffer, format='JPEG', quality=85)
        img_buffer.seek(0)

        return Image(img_buffer, width=new_width, height=new_height)

    # Estilos customizados
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.HexColor('#8B0000'),
        alignment=TA_LEFT,
        spaceAfter=2,
        leading=16
    )

    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.gray,
        alignment=TA_LEFT,
        spaceAfter=0
    )

    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=4
    )

    # ============================================================
    # CABEÇALHO COM LOGO
    # ============================================================
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_cbmgo.png')

    # Criar elementos do cabeçalho
    titulo_principal = Paragraph("CORPO DE BOMBEIROS MILITAR<br/>DO ESTADO DE GOIÁS", title_style)
    subtitulo = Paragraph("Sistema de Gestão de Missões - SIGEM", subtitle_style)

    # Verificar se a logo existe
    if os.path.exists(logo_path):
        try:
            # Logo com proporção mantida (max 2cm x 2cm) - preserva transparência
            logo = get_image_with_aspect_ratio(logo_path, 2*cm, 2*cm, preserve_transparency=True)
            # Tabela: Logo à esquerda, Título à direita
            header_data = [[logo, [titulo_principal, subtitulo]]]
            header_table = Table(header_data, colWidths=[2.5*cm, 14.5*cm])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('ALIGN', (1, 0), (1, 0), 'LEFT'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))
        except Exception:
            # Se der erro na logo, título centralizado
            title_style.alignment = TA_CENTER
            subtitle_style.alignment = TA_CENTER
            header_data = [[[titulo_principal, subtitulo]]]
            header_table = Table(header_data, colWidths=[17*cm])
            header_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ]))
    else:
        # Sem logo - apenas título centralizado
        title_style.alignment = TA_CENTER
        subtitle_style.alignment = TA_CENTER
        header_data = [[[titulo_principal, subtitulo]]]
        header_table = Table(header_data, colWidths=[17*cm])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))

    elements.append(header_table)
    elements.append(Spacer(1, 0.5*cm))

    # Linha separadora
    linha_sep = Table([['']], colWidths=[17*cm], rowHeights=[2])
    linha_sep.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#8B0000')),
    ]))
    elements.append(linha_sep)
    elements.append(Spacer(1, 0.5*cm))

    # ============================================================
    # TÍTULO DO RELATÓRIO
    # ============================================================
    elements.append(Paragraph(
        "<b>RELATÓRIO DE DESIGNAÇÕES</b>",
        ParagraphStyle('ReportTitle', fontSize=12, alignment=TA_CENTER, spaceAfter=15, textColor=colors.HexColor('#8B0000'))
    ))

    # ============================================================
    # DADOS DO OFICIAL COM FOTO
    # ============================================================
    # Verificar se oficial tem foto
    foto_path = None
    if oficial.foto:
        foto_path = oficial.foto.path if hasattr(oficial.foto, 'path') else None

    # Se não tem foto, usar avatar padrão
    if not foto_path or not os.path.exists(foto_path):
        foto_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'default_avatar.png')

    # Informações do oficial
    info_oficial = []
    info_oficial.append(Paragraph(f"<b>{oficial.posto} {oficial.nome}</b>",
                                   ParagraphStyle('OficialNome', fontSize=11, spaceAfter=4)))
    info_oficial.append(Paragraph(f"<b>RG:</b> {oficial.rg}", info_style))
    info_oficial.append(Paragraph(f"<b>Quadro:</b> {oficial.quadro}", info_style))
    info_oficial.append(Paragraph(f"<b>OBM:</b> {oficial.obm or 'Não informado'}", info_style))
    info_oficial.append(Paragraph(f"<b>Data do Relatório:</b> {datetime.now().strftime('%d/%m/%Y às %H:%M')}", info_style))

    # Criar tabela com foto e informações
    if os.path.exists(foto_path):
        try:
            # Foto com proporção mantida (max 2.5cm x 3cm)
            foto_oficial = get_image_with_aspect_ratio(foto_path, 2.5*cm, 3*cm, preserve_transparency=False)
            oficial_data = [[foto_oficial, info_oficial]]
            oficial_table = Table(oficial_data, colWidths=[3*cm, 14*cm])
            oficial_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))
            elements.append(oficial_table)
        except Exception as e:
            # Se der erro na foto, mostrar só as informações
            for info in info_oficial:
                elements.append(info)
    else:
        for info in info_oficial:
            elements.append(info)

    elements.append(Spacer(1, 0.5*cm))

    # Resumo
    from django.db.models import F
    total_ativas = designacoes.filter(missao__status='EM_ANDAMENTO').count()
    designacoes_ativas_anotadas = designacoes.filter(missao__status='EM_ANDAMENTO').annotate(
        soma=F('funcao__tde') + F('funcao__nqt') + F('funcao__grs') + F('funcao__dec')
    )
    total_baixa = designacoes_ativas_anotadas.filter(soma__gte=4, soma__lte=6).count()
    total_media = designacoes_ativas_anotadas.filter(soma__gte=7, soma__lte=9).count()
    total_alta = designacoes_ativas_anotadas.filter(soma__gte=10, soma__lte=12).count()

    resumo_data = [
        ['RESUMO DE MISSÕES EM ANDAMENTO', '', '', ''],
        ['Total Ativas', 'Baixa Complexidade', 'Média Complexidade', 'Alta Complexidade'],
        [str(total_ativas), str(total_baixa), str(total_media), str(total_alta)],
    ]

    resumo_table = Table(resumo_data, colWidths=[4*cm, 4*cm, 4*cm, 4*cm])
    resumo_table.setStyle(TableStyle([
        ('SPAN', (0, 0), (-1, 0)),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B0000')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f3f4f6')),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(resumo_table)
    elements.append(Spacer(1, 0.7*cm))

    # Tabela de designações
    if designacoes.exists():
        elements.append(Paragraph("<b>DETALHAMENTO DAS DESIGNAÇÕES</b>", ParagraphStyle('Heading', fontSize=11, spaceAfter=10)))

        table_data = [['Missão', 'Função', 'Complexidade', 'Status', 'Período']]

        for d in designacoes:
            periodo = ''
            if d.missao.data_inicio:
                periodo = d.missao.data_inicio.strftime('%d/%m/%Y')
                if d.missao.data_fim:
                    periodo += f" - {d.missao.data_fim.strftime('%d/%m/%Y')}"

            table_data.append([
                d.missao.nome[:40] + '...' if len(d.missao.nome) > 40 else d.missao.nome,
                d.funcao.funcao,
                d.get_complexidade_display(),
                d.missao.get_status_display(),
                periodo or '-'
            ])

        designacoes_table = Table(table_data, colWidths=[6*cm, 3*cm, 2.5*cm, 2.5*cm, 3*cm])
        designacoes_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B0000')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]))
        elements.append(designacoes_table)
    else:
        elements.append(Paragraph("Nenhuma designação encontrada.", info_style))

    # Rodapé
    elements.append(Spacer(1, 1*cm))
    elements.append(Paragraph(
        f"Documento gerado pelo SIGEM em {datetime.now().strftime('%d/%m/%Y às %H:%M')}",
        ParagraphStyle('Footer', fontSize=8, textColor=colors.gray, alignment=TA_CENTER)
    ))

    # Gerar PDF
    doc.build(elements)

    # Retornar resposta
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=relatorio_designacoes_{oficial.rg}.pdf'

    return response


# ============================================================
# 📤 IMPORTAÇÃO EM MASSA
# ============================================================
@login_required
@require_POST
def importar_excel(request, tipo):
    """Importa dados de arquivo Excel."""

    if not request.user.is_admin:
        messages.error(request, 'Sem permissão.')
        return redirect('admin_painel')

    import openpyxl
    from datetime import datetime

    arquivo = request.FILES.get('arquivo')

    if not arquivo:
        messages.error(request, 'Nenhum arquivo enviado.')
        return redirect('admin_painel')

    try:
        wb = openpyxl.load_workbook(arquivo)
        ws = wb.active

        count = 0
        errors = []

        # ============================================================
        # IMPORTAR OFICIAIS
        # ============================================================
        if tipo == 'oficiais':
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0]:  # Se tem CPF
                    try:
                        cpf = str(row[0]).replace('.', '').replace('-', '').strip()
                        oficial, created = Oficial.objects.update_or_create(
                            cpf=cpf,
                            defaults={
                                'rg': str(row[1]).strip() if row[1] else '',
                                'nome': str(row[2]).strip() if row[2] else '',
                                'nome_guerra': str(row[3]).strip() if row[3] else '',
                                'posto': str(row[4]).strip() if row[4] else '',
                                'quadro': str(row[5]).strip() if row[5] else '',
                                'obm': str(row[6]).strip() if row[6] else '',
                                'funcao': str(row[7]).strip() if row[7] else '',
                                'email': str(row[8]).strip() if row[8] else '',
                                'telefone': str(row[9]).strip() if row[9] else '',
                            }
                        )
                        count += 1

                        # Criar usuário automaticamente se não existir
                        if created and not Usuario.objects.filter(cpf=cpf).exists():
                            Usuario.objects.create_user(
                                cpf=cpf,
                                password='123456',
                                oficial=oficial,
                                role='oficial'
                            )
                    except Exception as e:
                        errors.append(f'Linha {row_num}: {str(e)}')

        # ============================================================
        # IMPORTAR MISSÕES
        # ============================================================
        elif tipo == 'missoes':
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] and row[1]:  # Se tem tipo e nome
                    try:
                        # Processar datas
                        data_inicio = None
                        data_fim = None

                        if row[4]:
                            if isinstance(row[4], datetime):
                                data_inicio = row[4].date()
                            else:
                                data_inicio = datetime.strptime(str(row[4]), '%Y-%m-%d').date()

                        if row[5]:
                            if isinstance(row[5], datetime):
                                data_fim = row[5].date()
                            else:
                                data_fim = datetime.strptime(str(row[5]), '%Y-%m-%d').date()

                        Missao.objects.create(
                            tipo=str(row[0]).strip().upper(),
                            nome=str(row[1]).strip(),
                            descricao=str(row[2]).strip() if row[2] else '',
                            local=str(row[3]).strip() if row[3] else '',
                            data_inicio=data_inicio,
                            data_fim=data_fim,
                            status=str(row[6]).strip().upper() if row[6] else 'PLANEJADA',
                            documento_referencia=str(row[7]).strip() if row[7] else '',
                        )
                        count += 1
                    except Exception as e:
                        errors.append(f'Linha {row_num}: {str(e)}')

        # ============================================================
        # IMPORTAR DESIGNAÇÕES
        # ============================================================
        elif tipo == 'designacoes':
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] and row[1]:  # Se tem missao_id e oficial_rg
                    try:
                        missao_id = int(row[0])
                        oficial_rg = str(row[1]).strip()

                        # Buscar missão e oficial
                        from ..models import Funcao
                        missao = Missao.objects.get(id=missao_id)
                        oficial = Oficial.objects.get(rg=oficial_rg)

                        # Buscar ou criar função (usar nome da coluna 2)
                        funcao_nome = str(row[2]).strip().upper() if row[2] else 'MEMBRO'
                        funcao, created = Funcao.objects.get_or_create(
                            missao=missao,
                            funcao=funcao_nome,
                            defaults={
                                'tde': 2,  # Medium default
                                'nqt': 2,
                                'grs': 2,
                                'dec': 2
                            }
                        )

                        Designacao.objects.update_or_create(
                            missao=missao,
                            oficial=oficial,
                            defaults={
                                'funcao': funcao,
                                'observacoes': str(row[4]).strip() if row[4] else '',
                            }
                        )
                        count += 1
                    except Missao.DoesNotExist:
                        errors.append(f'Linha {row_num}: Missão ID {row[0]} não encontrada')
                    except Oficial.DoesNotExist:
                        errors.append(f'Linha {row_num}: Oficial RG {row[1]} não encontrado')
                    except Exception as e:
                        errors.append(f'Linha {row_num}: {str(e)}')

        # ============================================================
        # IMPORTAR UNIDADES
        # ============================================================
        elif tipo == 'unidades':
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0]:  # Se tem nome
                    try:
                        comando_superior = None
                        if row[3]:
                            try:
                                comando_superior = Unidade.objects.get(id=int(row[3]))
                            except Unidade.DoesNotExist:
                                pass

                        Unidade.objects.update_or_create(
                            nome=str(row[0]).strip(),
                            defaults={
                                'sigla': str(row[1]).strip() if row[1] else '',
                                'tipo': str(row[2]).strip().upper() if row[2] else '',
                                'comando_superior': comando_superior,
                            }
                        )
                        count += 1
                    except Exception as e:
                        errors.append(f'Linha {row_num}: {str(e)}')

        # ============================================================
        # IMPORTAR USUÁRIOS
        # ============================================================
        elif tipo == 'usuarios':
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0]:  # Se tem CPF
                    try:
                        cpf = str(row[0]).replace('.', '').replace('-', '').strip()
                        role = str(row[1]).strip().lower() if row[1] else 'oficial'

                        # Buscar oficial vinculado pelo RG
                        oficial = None
                        if row[2]:
                            try:
                                oficial = Oficial.objects.get(rg=str(row[2]).strip())
                            except Oficial.DoesNotExist:
                                pass

                        if not Usuario.objects.filter(cpf=cpf).exists():
                            Usuario.objects.create_user(
                                cpf=cpf,
                                password='123456',
                                role=role,
                                oficial=oficial,
                            )
                            count += 1
                        else:
                            # Atualizar usuário existente
                            usuario = Usuario.objects.get(cpf=cpf)
                            usuario.role = role
                            if oficial:
                                usuario.oficial = oficial
                            usuario.save()
                            count += 1
                    except Exception as e:
                        errors.append(f'Linha {row_num}: {str(e)}')

        # Mensagem de resultado
        if count > 0:
            messages.success(request, f'{count} registros importados/atualizados com sucesso!')

        if errors:
            error_msg = f'Erros encontrados ({len(errors)}): ' + '; '.join(errors[:5])
            if len(errors) > 5:
                error_msg += f' ... e mais {len(errors) - 5} erros.'
            messages.warning(request, error_msg)

    except Exception as e:
        messages.error(request, f'Erro na importação: {str(e)}')

    return redirect('admin_painel')