"""
============================================================
🎯 SIGEM - Views (Lógica das Páginas)
Sistema de Gestão de Missões - CBMGO
============================================================
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, Q, Avg
from django.utils import timezone
from django.views.decorators.http import require_POST, require_GET
from django.core.paginator import Paginator

from .models import Oficial, Missao, Designacao, Unidade, Usuario, SolicitacaoDesignacao, SolicitacaoMissao
from .decorators import (
    acesso_dashboard, acesso_comparar, acesso_admin_painel,
    permissao_gerenciar_oficiais, permissao_gerenciar_missoes,
    permissao_gerenciar_designacoes, permissao_gerenciar_unidades,
    permissao_gerenciar_usuarios, permissao_gerenciar_solicitacoes
)


# ============================================================
# 🔐 AUTENTICAÇÃO
# ============================================================
def login_view(request):
    """Página de login."""
    if request.user.is_authenticated:
        return redirect('redirecionar_por_perfil')
    
    if request.method == 'POST':
        cpf = request.POST.get('cpf', '').replace('.', '').replace('-', '')
        senha = request.POST.get('senha', '')
        
        user = authenticate(request, cpf=cpf, password=senha)
        
        if user is not None:
            login(request, user)
            user.ultimo_acesso = timezone.now()
            user.save(update_fields=['ultimo_acesso'])
            messages.success(request, f'Bem-vindo, {user}!')
            return redirect('redirecionar_por_perfil')
        else:
            messages.error(request, 'CPF ou senha incorretos.')
    
    return render(request, 'auth/login.html')


@login_required
def redirecionar_por_perfil(request):
    """Redireciona o usuário para a página inicial conforme seu perfil."""
    user = request.user
    
    if user.role in ['admin', 'comando_geral']:
        return redirect('dashboard')
    elif user.role in ['corregedor', 'bm3', 'comandante']:
        return redirect('comparar_oficiais')
    else:  # oficial
        return redirect('painel_oficial')


@login_required
def logout_view(request):
    """Logout do usuário."""
    logout(request)
    messages.info(request, 'Você saiu do sistema.')
    return redirect('login')


# ============================================================
# 📊 DASHBOARD - VISÃO GERAL (Executivo)
# ============================================================
@login_required
@acesso_dashboard
def dashboard(request):
    """Dashboard executivo - Visão Geral para Comando-Geral e Comandantes."""
    
    from django.db.models import Sum, Case, When, IntegerField, F, Value
    from django.db.models.functions import TruncMonth, Coalesce
    from datetime import timedelta
    from collections import defaultdict
    import json
    import traceback
    
    # Em caso de erro, mostrar página de erro amigável
    try:
        hoje = timezone.now().date()
        user = request.user
        
        # ============================================================
        # 🔒 FILTRO POR OBM (COMANDANTE)
        # ============================================================
        is_comandante = user.is_comandante
        obms_permitidas = []
        obm_sigla = None
        
        if is_comandante:
            obms_permitidas = user.get_obm_subordinadas()
            if user.oficial and user.oficial.obm:
                obm_sigla = user.oficial.obm
            
            # QuerySets base filtrados por OBM
            oficiais_base = Oficial.objects.filter(ativo=True, obm__in=obms_permitidas)
            designacoes_base = Designacao.objects.filter(oficial__obm__in=obms_permitidas)
            
            # Missões onde há oficiais da OBM designados
            missoes_ids_com_oficiais_obm = designacoes_base.values_list('missao_id', flat=True).distinct()
            missoes_base = Missao.objects.filter(id__in=missoes_ids_com_oficiais_obm)
        else:
            # Visão total (admin, comando_geral)
            oficiais_base = Oficial.objects.filter(ativo=True)
            designacoes_base = Designacao.objects.all()
            missoes_base = Missao.objects.all()
        
        # ============================================================
        # 📌 KPIs PRINCIPAIS
        # ============================================================
        
        # Total de oficiais ativos
        total_oficiais = oficiais_base.count() or 0
        
        # Oficiais com pelo menos uma missão em andamento
        oficiais_com_missao = oficiais_base.filter(
            designacoes__missao__status='EM_ANDAMENTO'
        ).distinct().count() or 0
        
        # Taxa de ocupação
        taxa_ocupacao = round((oficiais_com_missao / total_oficiais * 100), 1) if total_oficiais > 0 else 0
        
        # Total de missões ativas (com oficiais da OBM para comandante)
        total_missoes_ativas = missoes_base.filter(status='EM_ANDAMENTO').count() or 0
        
        # Total de designações ativas
        total_designacoes_ativas = designacoes_base.filter(missao__status='EM_ANDAMENTO').count() or 0
        
        # Carga média por oficial (apenas os que têm missão)
        carga_media = round(total_designacoes_ativas / oficiais_com_missao, 1) if oficiais_com_missao > 0 else 0
        
        # Designações por complexidade (ativas)
        designacoes_baixa = designacoes_base.filter(missao__status='EM_ANDAMENTO', complexidade='BAIXA').count() or 0
        designacoes_media = designacoes_base.filter(missao__status='EM_ANDAMENTO', complexidade='MEDIA').count() or 0
        designacoes_alta = designacoes_base.filter(missao__status='EM_ANDAMENTO', complexidade='ALTA').count() or 0
        
        # Índice de complexidade alta
        indice_alta = round((designacoes_alta / total_designacoes_ativas * 100), 1) if total_designacoes_ativas > 0 else 0
        
        # Solicitações pendentes (apenas para não-comandantes)
        if not is_comandante:
            solicitacoes_pendentes = (
                SolicitacaoDesignacao.objects.filter(status='PENDENTE').count() +
                SolicitacaoMissao.objects.filter(status='PENDENTE').count()
            ) or 0
        else:
            solicitacoes_pendentes = 0
        
        # Taxa de conclusão (missões concluídas / total não canceladas)
        total_missoes_nao_canceladas = missoes_base.exclude(status='CANCELADA').count() or 0
        missoes_concluidas = missoes_base.filter(status='CONCLUIDA').count() or 0
        taxa_conclusao = round((missoes_concluidas / total_missoes_nao_canceladas * 100), 1) if total_missoes_nao_canceladas > 0 else 0
        
        # ============================================================
        # 📈 EVOLUÇÃO MENSAL (últimos 12 meses)
        # ============================================================
        
        doze_meses_atras = hoje - timedelta(days=365)
        
        evolucao_mensal = missoes_base.filter(
            criado_em__date__gte=doze_meses_atras
        ).annotate(
            mes=TruncMonth('criado_em')
        ).values('mes').annotate(
            criadas=Count('id'),
            em_andamento=Count('id', filter=Q(status='EM_ANDAMENTO')),
            concluidas=Count('id', filter=Q(status='CONCLUIDA'))
        ).order_by('mes')
        
        # Formatar para Chart.js
        evolucao_labels = []
        evolucao_criadas = []
        evolucao_andamento = []
        evolucao_concluidas = []
        
        meses_nome = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
        for item in evolucao_mensal:
            if item['mes']:
                evolucao_labels.append(f"{meses_nome[item['mes'].month - 1]}/{str(item['mes'].year)[2:]}")
                evolucao_criadas.append(item['criadas'])
                evolucao_andamento.append(item['em_andamento'])
                evolucao_concluidas.append(item['concluidas'])
        
        # ============================================================
        # 🥧 MISSÕES POR TIPO
        # ============================================================
        
        missoes_por_tipo = missoes_base.filter(
            status='EM_ANDAMENTO'
        ).values('tipo').annotate(
            total=Count('id')
        ).order_by('-total')
        
        tipo_labels = []
        tipo_valores = []
        tipo_display = dict(Missao.TIPO_CHOICES)
        
        for item in missoes_por_tipo:
            tipo_labels.append(tipo_display.get(item['tipo'], item['tipo']))
            tipo_valores.append(item['total'])
        
        # ============================================================
        # 📊 CARGA POR OBM
        # ============================================================
        
        # Agregar dados por OBM (filtrado para comandante)
        if is_comandante:
            oficiais_por_obm = oficiais_base.exclude(
                Q(obm__isnull=True) | Q(obm='')
            ).values('obm').annotate(
                efetivo=Count('id'),
                em_missao=Count('id', filter=Q(designacoes__missao__status='EM_ANDAMENTO'), distinct=True),
            ).order_by('-efetivo')
        else:
            oficiais_por_obm = Oficial.objects.filter(ativo=True).exclude(
                Q(obm__isnull=True) | Q(obm='')
            ).values('obm').annotate(
                efetivo=Count('id'),
                em_missao=Count('id', filter=Q(designacoes__missao__status='EM_ANDAMENTO'), distinct=True),
            ).order_by('-efetivo')[:10]
        
        # Calcular carga por complexidade para cada OBM
        carga_por_obm = []
        for obm_data in oficiais_por_obm:
            obm_nome = obm_data['obm']
            
            # Contar designações por complexidade
            baixa = Designacao.objects.filter(
                oficial__obm=obm_nome,
                missao__status='EM_ANDAMENTO',
                complexidade='BAIXA'
            ).count()
            media = Designacao.objects.filter(
                oficial__obm=obm_nome,
                missao__status='EM_ANDAMENTO',
                complexidade='MEDIA'
            ).count()
            alta = Designacao.objects.filter(
                oficial__obm=obm_nome,
                missao__status='EM_ANDAMENTO',
                complexidade='ALTA'
            ).count()
            
            carga_por_obm.append({
                'obm': obm_nome,
                'efetivo': obm_data['efetivo'],
                'em_missao': obm_data['em_missao'],
                'disponivel': obm_data['efetivo'] - obm_data['em_missao'],
                'baixa': baixa,
                'media': media,
                'alta': alta,
                'carga_total': baixa + (media * 2) + (alta * 3),
                'ocupacao': round((obm_data['em_missao'] / obm_data['efetivo'] * 100), 0) if obm_data['efetivo'] > 0 else 0
            })
        
        # Ordenar por carga total
        carga_por_obm.sort(key=lambda x: x['carga_total'], reverse=True)
        
        obm_labels = [item['obm'][:15] for item in carga_por_obm]
        obm_baixa = [item['baixa'] for item in carga_por_obm]
        obm_media = [item['media'] for item in carga_por_obm]
        obm_alta = [item['alta'] for item in carga_por_obm]
        
        # ============================================================
        # 🏆 TOP 10 OFICIAIS COM MAIOR CARGA
        # ============================================================
        
        oficiais_top = oficiais_base.annotate(
            total_missoes=Count('designacoes', filter=Q(designacoes__missao__status='EM_ANDAMENTO')),
            qtd_baixa=Count('designacoes', filter=Q(designacoes__missao__status='EM_ANDAMENTO', designacoes__complexidade='BAIXA')),
            qtd_media=Count('designacoes', filter=Q(designacoes__missao__status='EM_ANDAMENTO', designacoes__complexidade='MEDIA')),
            qtd_alta=Count('designacoes', filter=Q(designacoes__missao__status='EM_ANDAMENTO', designacoes__complexidade='ALTA')),
            qtd_chefia=Count('designacoes', filter=Q(
                designacoes__missao__status='EM_ANDAMENTO',
                designacoes__funcao_na_missao__in=['COMANDANTE', 'SUBCOMANDANTE', 'COORDENADOR', 'PRESIDENTE', 'ENCARREGADO']
            ))
        ).annotate(
            carga_ponderada=F('qtd_baixa') + (F('qtd_media') * 2) + (F('qtd_alta') * 3)
        ).filter(total_missoes__gt=0).order_by('-carga_ponderada')[:50]
        
        # ============================================================
        # ⚠️ ALERTAS DO SISTEMA
        # ============================================================
        
        alertas = []
        
        # 🔴 Oficiais com sobrecarga (carga > 20)
        oficiais_sobrecarga = oficiais_base.annotate(
            carga=Count('designacoes', filter=Q(designacoes__missao__status='EM_ANDAMENTO', designacoes__complexidade='BAIXA')) +
                  Count('designacoes', filter=Q(designacoes__missao__status='EM_ANDAMENTO', designacoes__complexidade='MEDIA')) * 2 +
                  Count('designacoes', filter=Q(designacoes__missao__status='EM_ANDAMENTO', designacoes__complexidade='ALTA')) * 3
        ).filter(carga__gt=20).count()
        
        if oficiais_sobrecarga > 0:
            alertas.append({
                'nivel': 'critico',
                'icone': 'alert-triangle',
                'mensagem': f'{oficiais_sobrecarga} oficial(is) com sobrecarga de trabalho',
                'descricao': 'Carga ponderada superior a 20 pontos'
            })
        
        # 🔴 OBMs com ocupação > 90%
        obms_sobrecarga = [obm for obm in carga_por_obm if obm['ocupacao'] > 90]
        if obms_sobrecarga:
            alertas.append({
                'nivel': 'critico',
                'icone': 'building',
                'mensagem': f'{len(obms_sobrecarga)} OBM(s) com ocupação acima de 90%',
                'descricao': ', '.join([o['obm'] for o in obms_sobrecarga[:3]])
            })
        
        # 🟠 Missões sem designação (apenas para não-comandantes)
        if not is_comandante:
            missoes_sem_designacao = Missao.objects.filter(
                status='EM_ANDAMENTO'
            ).annotate(
                total_designados=Count('designacoes')
            ).filter(total_designados=0).count()
            
            if missoes_sem_designacao > 0:
                alertas.append({
                    'nivel': 'alto',
                    'icone': 'users',
                    'mensagem': f'{missoes_sem_designacao} missão(ões) sem oficiais designados',
                    'descricao': 'Missões em andamento sem nenhum responsável'
                })
            
            # 🟠 Solicitações pendentes há mais de 7 dias
            sete_dias_atras = timezone.now() - timedelta(days=7)
            solicitacoes_atrasadas = (
                SolicitacaoDesignacao.objects.filter(
                    status='PENDENTE',
                    criado_em__lt=sete_dias_atras
                ).count() +
                SolicitacaoMissao.objects.filter(
                    status='PENDENTE',
                    criado_em__lt=sete_dias_atras
                ).count()
            )
            
            if solicitacoes_atrasadas > 0:
                alertas.append({
                    'nivel': 'alto',
                    'icone': 'clock',
                    'mensagem': f'{solicitacoes_atrasadas} solicitação(ões) pendente(s) há mais de 7 dias',
                    'descricao': 'Necessitam avaliação urgente'
                })
        
        # 🟡 Oficiais sem missão
        oficiais_sem_missao = total_oficiais - oficiais_com_missao
        if oficiais_sem_missao > 0 and total_oficiais > 0 and (oficiais_sem_missao / total_oficiais) > 0.3:
            alertas.append({
                'nivel': 'medio',
                'icone': 'user-x',
                'mensagem': f'{oficiais_sem_missao} oficial(is) sem missão atribuída',
                'descricao': f'Representa {round((oficiais_sem_missao/total_oficiais)*100)}% do efetivo'
            })
        
        # 🟡 Missões próximas do prazo (7 dias)
        proxima_semana = hoje + timedelta(days=7)
        missoes_prazo = missoes_base.filter(
            status='EM_ANDAMENTO',
            data_fim__lte=proxima_semana,
            data_fim__gte=hoje
        ).count()
        
        if missoes_prazo > 0:
            alertas.append({
                'nivel': 'medio',
                'icone': 'calendar',
                'mensagem': f'{missoes_prazo} missão(ões) com prazo nos próximos 7 dias',
                'descricao': 'Acompanhar conclusão'
            })
        
        # ============================================================
        # 📋 MISSÕES RECENTES
        # ============================================================
        
        missoes_recentes = missoes_base.select_related().annotate(
            qtd_designados=Count('designacoes')
        ).order_by('-criado_em')[:5]
        
        # ============================================================
        # 👔 DISTRIBUIÇÃO POR POSTO
        # ============================================================
        
        distribuicao_posto = oficiais_base.values('posto').annotate(
            efetivo=Count('id'),
            em_missao=Count('id', filter=Q(designacoes__missao__status='EM_ANDAMENTO'), distinct=True),
            total_designacoes=Count('designacoes', filter=Q(designacoes__missao__status='EM_ANDAMENTO')),
            qtd_chefia=Count('designacoes', filter=Q(
                designacoes__missao__status='EM_ANDAMENTO',
                designacoes__funcao_na_missao__in=['COMANDANTE', 'SUBCOMANDANTE', 'COORDENADOR', 'PRESIDENTE', 'ENCARREGADO']
            ))
        ).order_by('posto')
        
        posto_ordem = ['Cel', 'TC', 'Maj', 'Cap', '1º Ten', '2º Ten', 'Asp']
        posto_dict = {p['posto']: p for p in distribuicao_posto}
        distribuicao_posto_ordenada = []
        
        for posto in posto_ordem:
            if posto in posto_dict:
                p = posto_dict[posto]
                p['carga_media'] = round(p['total_designacoes'] / p['em_missao'], 1) if p['em_missao'] > 0 else 0
                p['perc_chefia'] = round((p['qtd_chefia'] / p['total_designacoes'] * 100), 0) if p['total_designacoes'] > 0 else 0
                distribuicao_posto_ordenada.append(p)
        
        posto_labels = [p['posto'] for p in distribuicao_posto_ordenada]
        posto_efetivo = [p['efetivo'] for p in distribuicao_posto_ordenada]
        posto_em_missao = [p['em_missao'] for p in distribuicao_posto_ordenada]
        
        # ============================================================
        # 📊 DISTRIBUIÇÃO POR QUADRO
        # ============================================================
        
        distribuicao_quadro = oficiais_base.values('quadro').annotate(
            efetivo=Count('id'),
            em_missao=Count('id', filter=Q(designacoes__missao__status='EM_ANDAMENTO'), distinct=True)
        ).order_by('-efetivo')
        
        quadro_labels = [q['quadro'] for q in distribuicao_quadro]
        quadro_valores = [q['efetivo'] for q in distribuicao_quadro]
        
        # ============================================================
        # 📅 DADOS PARA ABA TEMPORAL
        # ============================================================
        
        # Duração média por tipo de missão (apenas concluídas)
        duracao_por_tipo = []
        for tipo_code, tipo_nome in Missao.TIPO_CHOICES:
            missoes_tipo = missoes_base.filter(
                tipo=tipo_code,
                status='CONCLUIDA',
                data_inicio__isnull=False,
                data_fim__isnull=False
            )
            if missoes_tipo.exists():
                total_dias = 0
                count = 0
                for m in missoes_tipo:
                    if m.data_fim and m.data_inicio:
                        total_dias += (m.data_fim - m.data_inicio).days
                        count += 1
                if count > 0:
                    duracao_por_tipo.append({
                        'tipo': tipo_nome,
                        'duracao_media': round(total_dias / count, 0)
                    })
        
        duracao_por_tipo.sort(key=lambda x: x['duracao_media'], reverse=True)
        
        # ============================================================
        # CONTEXTO FINAL
        # ============================================================
        
        # Calcular percentuais de complexidade
        perc_baixa = round((designacoes_baixa / total_designacoes_ativas * 100), 0) if total_designacoes_ativas > 0 else 0
        perc_media = round((designacoes_media / total_designacoes_ativas * 100), 0) if total_designacoes_ativas > 0 else 0
        perc_alta = round((designacoes_alta / total_designacoes_ativas * 100), 0) if total_designacoes_ativas > 0 else 0
        
        context = {
            # Identificação do perfil
            'is_comandante': is_comandante,
            'obm_sigla': obm_sigla,
            
            # KPIs
            'total_oficiais': total_oficiais,
            'oficiais_com_missao': oficiais_com_missao,
            'taxa_ocupacao': taxa_ocupacao,
            'total_missoes_ativas': total_missoes_ativas,
            'total_designacoes_ativas': total_designacoes_ativas,
            'carga_media': carga_media,
            'designacoes_baixa': designacoes_baixa,
            'designacoes_media': designacoes_media,
            'designacoes_alta': designacoes_alta,
            'perc_baixa': perc_baixa,
            'perc_media': perc_media,
            'perc_alta': perc_alta,
            'indice_alta': indice_alta,
            'solicitacoes_pendentes': solicitacoes_pendentes,
            'taxa_conclusao': taxa_conclusao,
            'missoes_concluidas': missoes_concluidas,
            
            # Evolução mensal (JSON para Chart.js)
            'evolucao_labels': json.dumps(evolucao_labels),
            'evolucao_criadas': json.dumps(evolucao_criadas),
            'evolucao_andamento': json.dumps(evolucao_andamento),
            'evolucao_concluidas': json.dumps(evolucao_concluidas),
            
            # Missões por tipo
            'tipo_labels': json.dumps(tipo_labels),
            'tipo_valores': json.dumps(tipo_valores),
            
            # Carga por OBM
            'carga_por_obm': carga_por_obm,
            'obm_labels': json.dumps(obm_labels),
            'obm_baixa': json.dumps(obm_baixa),
            'obm_media': json.dumps(obm_media),
            'obm_alta': json.dumps(obm_alta),
            
            # Top oficiais
            'oficiais_top': oficiais_top,
            
            # Alertas
            'alertas': alertas,
            
            # Missões recentes
            'missoes_recentes': missoes_recentes,
            
            # Distribuição por posto
            'distribuicao_posto': distribuicao_posto_ordenada,
            'posto_labels': json.dumps(posto_labels),
            'posto_efetivo': json.dumps(posto_efetivo),
            'posto_em_missao': json.dumps(posto_em_missao),
            
            # Distribuição por quadro
            'distribuicao_quadro': distribuicao_quadro,
            'quadro_labels': json.dumps(quadro_labels),
            'quadro_valores': json.dumps(quadro_valores),
            
            # Duração por tipo
            'duracao_por_tipo': duracao_por_tipo,
            
            # Oficiais sem missão
            'oficiais_sem_missao': total_oficiais - oficiais_com_missao,
        }
        
        return render(request, 'pages/dashboard.html', context)
    
    except Exception as e:
        # Em caso de erro, mostrar página com informação do erro
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Erro no dashboard: {str(e)}\n{traceback.format_exc()}")
        
        # Retornar página de erro amigável com detalhes para debug
        error_context = {
            'error_message': str(e),
            'error_type': type(e).__name__,
        }
        return render(request, 'pages/dashboard_error.html', error_context)


# ============================================================
# ⚖️ COMPARAR OFICIAIS
# ============================================================
@login_required
@acesso_comparar
def comparar_oficiais(request):
    """Página para comparar carga de trabalho entre oficiais."""
    
    user = request.user
    
    # Filtros disponíveis
    postos = Oficial.POSTO_CHOICES
    quadros = Oficial.QUADRO_CHOICES
    
    # Para comandante, filtrar apenas OBMs permitidas
    if user.is_comandante:
        obms_permitidas = user.get_obm_subordinadas()
        obms = obms_permitidas
    else:
        obms = Oficial.objects.values_list('obm', flat=True).distinct().order_by('obm')
    
    context = {
        'postos': postos,
        'quadros': quadros,
        'obms': obms,
        'is_comandante': user.is_comandante,
    }
    
    return render(request, 'pages/comparar_oficiais.html', context)


# ============================================================
# 🗂️ DASHBOARD DE MISSÕES
# ============================================================
@login_required
def missoes_dashboard(request):
    """Dashboard completo de missões com organograma."""
    
    # Totalizadores por tipo (em andamento)
    tipos = ['OPERACIONAL', 'ADMINISTRATIVA', 'ENSINO', 'CORREICIONAL', 'COMISSAO', 'ACAO_SOCIAL']
    
    totais_por_tipo = []
    for tipo in tipos:
        total = Missao.objects.filter(tipo=tipo, status='EM_ANDAMENTO').count()
        totais_por_tipo.append({
            'tipo': tipo,
            'tipo_display': dict(Missao.TIPO_CHOICES).get(tipo, tipo),
            'total': total
        })
    
    # Filtros disponíveis
    status_choices = Missao.STATUS_CHOICES
    
    context = {
        'totais_por_tipo': totais_por_tipo,
        'status_choices': status_choices,
        'tipo_choices': Missao.TIPO_CHOICES,
    }
    
    return render(request, 'pages/missoes.html', context)


# ============================================================
# 👤 CONSULTAR OFICIAL (antigo Painel do Oficial)
# ============================================================
@login_required
def consultar_oficial(request, oficial_id=None):
    """
    Consulta painel de um oficial.
    - Oficial: vê apenas seu próprio painel
    - Comandante: vê oficiais da sua OBM e subordinadas
    - Admin/Comando-Geral: vê todos os oficiais
    """
    
    usuario = request.user
    oficial = None
    
    # Verificar se pode consultar outros oficiais
    pode_consultar_outros = usuario.role in ['admin', 'comando_geral', 'comandante']
    
    # Determinar qual oficial será exibido
    if oficial_id:
        # Tentando ver outro oficial
        oficial = get_object_or_404(Oficial, pk=oficial_id)
        
        # Verificar permissão
        if not usuario.pode_ver_oficial(oficial):
            messages.error(request, 'Você não tem permissão para visualizar este oficial.')
            return redirect('consultar_oficial')
    else:
        # Se pode consultar outros, mostrar apenas a tela de busca (sem oficial pré-selecionado)
        if pode_consultar_outros:
            # Admin/Comandante podem acessar sem ter oficial vinculado
            oficial = usuario.oficial  # Pode ser None
        else:
            # Perfil oficial: deve ver seu próprio painel
            oficial = usuario.oficial
            if not oficial:
                messages.warning(request, 'Seu usuário não está vinculado a um oficial.')
                return redirect('dashboard')
    
    # Lista de OBMs disponíveis para o filtro
    obms_disponiveis = []
    if pode_consultar_outros:
        if usuario.role in ['admin', 'comando_geral']:
            # Admin e Comando-Geral veem todas as OBMs
            obms_disponiveis = list(
                Oficial.objects.filter(ativo=True)
                .exclude(obm__isnull=True)
                .exclude(obm='')
                .values_list('obm', flat=True)
                .distinct()
                .order_by('obm')
            )
        elif usuario.role == 'comandante':
            # Comandante vê apenas sua OBM e subordinadas
            obms_disponiveis = usuario.get_obm_subordinadas()
    
    # Designações do oficial (se houver oficial selecionado)
    designacoes = []
    if oficial:
        designacoes = Designacao.objects.filter(oficial=oficial).select_related('missao').order_by('-criado_em')
        
        # Filtros
        tipo = request.GET.get('tipo', '')
        status = request.GET.get('status', '')
        complexidade = request.GET.get('complexidade', '')
        
        if tipo:
            designacoes = designacoes.filter(missao__tipo=tipo)
        if status:
            designacoes = designacoes.filter(missao__status=status)
        if complexidade:
            designacoes = designacoes.filter(complexidade=complexidade)
    
    # Verificar se está vendo o próprio painel
    visualizando_proprio = oficial and usuario.oficial and usuario.oficial.id == oficial.id
    
    context = {
        'oficial': oficial,
        'designacoes': designacoes,
        'tipo_choices': Missao.TIPO_CHOICES,
        'status_choices': Missao.STATUS_CHOICES,
        'complexidade_choices': Designacao.COMPLEXIDADE_CHOICES,
        'pode_consultar_outros': pode_consultar_outros,
        'obms_disponiveis': obms_disponiveis,
        'posto_choices': Oficial.POSTO_CHOICES,
        'quadro_choices': Oficial.QUADRO_CHOICES,
        'visualizando_proprio': visualizando_proprio,
    }
    
    return render(request, 'pages/consultar_oficial.html', context)


@login_required
def htmx_buscar_oficiais(request):
    """Busca oficiais para o painel de consulta (HTMX)."""
    
    usuario = request.user
    
    # Verificar permissão
    if usuario.role not in ['admin', 'comando_geral', 'comandante']:
        return HttpResponse('<p class="text-danger">Sem permissão.</p>')
    
    # Base query
    oficiais = Oficial.objects.filter(ativo=True)
    
    # Filtrar por permissão do comandante
    if usuario.role == 'comandante':
        obms_permitidas = usuario.get_obm_subordinadas()
        if obms_permitidas:
            q_filter = Q()
            for obm in obms_permitidas:
                q_filter |= Q(obm__icontains=obm)
            oficiais = oficiais.filter(q_filter)
    
    # Aplicar filtros da busca
    rg = request.GET.get('rg', '').strip()
    nome = request.GET.get('nome', '').strip()
    obm = request.GET.get('obm', '').strip()
    posto = request.GET.get('posto', '').strip()
    quadro = request.GET.get('quadro', '').strip()
    
    if rg:
        oficiais = oficiais.filter(rg__icontains=rg)
    if nome:
        oficiais = oficiais.filter(
            Q(nome__icontains=nome) | Q(nome_guerra__icontains=nome)
        )
    if obm:
        oficiais = oficiais.filter(obm__icontains=obm)
    if posto:
        oficiais = oficiais.filter(posto=posto)
    if quadro:
        oficiais = oficiais.filter(quadro=quadro)
    
    # Limitar resultados
    oficiais = oficiais.order_by('posto', 'nome')[:20]
    
    # Verificar se há filtros ativos
    tem_filtros = any([rg, nome, obm, posto, quadro])
    
    return render(request, 'htmx/oficiais_busca_resultado.html', {
        'oficiais': oficiais,
        'tem_filtros': tem_filtros,
        'total': oficiais.count(),
    })


# Alias para manter compatibilidade com URLs antigas
@login_required
def painel_oficial(request):
    """Painel do oficial logado com suas designações e formulários de solicitação."""
    
    usuario = request.user
    
    if not usuario.oficial:
        messages.error(request, 'Você não possui um oficial vinculado ao seu usuário.')
        return redirect('missoes_dashboard')
    
    oficial = usuario.oficial
    
    # Designações do oficial
    designacoes = Designacao.objects.filter(oficial=oficial).select_related('missao').order_by('-criado_em')
    
    # Filtros
    tipo = request.GET.get('tipo', '')
    status_filtro = request.GET.get('status', '')
    complexidade = request.GET.get('complexidade', '')
    
    if tipo:
        designacoes = designacoes.filter(missao__tipo=tipo)
    if status_filtro:
        designacoes = designacoes.filter(missao__status=status_filtro)
    if complexidade:
        designacoes = designacoes.filter(complexidade=complexidade)
    
    # Missões disponíveis para solicitar designação (Planejada ou Em Andamento)
    missoes_disponiveis = Missao.objects.filter(
        status__in=['PLANEJADA', 'EM_ANDAMENTO']
    ).order_by('nome')
    
    # Solicitações recentes (últimas 5 de cada tipo)
    solicitacoes_missao = SolicitacaoMissao.objects.filter(solicitante=oficial).order_by('-criado_em')[:5]
    solicitacoes_designacao = SolicitacaoDesignacao.objects.filter(solicitante=oficial).order_by('-criado_em')[:5]
    
    context = {
        'oficial': oficial,
        'designacoes': designacoes,
        'missoes_disponiveis': missoes_disponiveis,
        'solicitacoes_missao': solicitacoes_missao,
        'solicitacoes_designacao': solicitacoes_designacao,
        'tipo_choices': Missao.TIPO_CHOICES,
        'status_choices': Missao.STATUS_CHOICES,
        'complexidade_choices': Designacao.COMPLEXIDADE_CHOICES,
        'local_choices': SolicitacaoMissao.LOCAL_CHOICES,
    }
    
    return render(request, 'pages/painel_oficial.html', context)


# ============================================================
# 🔧 PAINEL ADMINISTRATIVO
# ============================================================
@login_required
@acesso_admin_painel
def admin_painel(request):
    """Painel administrativo com CRUD de todas as entidades."""
    
    user = request.user
    
    # Determinar aba inicial baseado nas permissões
    if user.is_admin:
        aba_padrao = 'oficiais'
    elif user.is_corregedor or user.is_bm3:
        aba_padrao = 'missoes'
    else:
        aba_padrao = 'missoes'
    
    context = {
        'aba_ativa': request.GET.get('aba', aba_padrao),
    }
    
    return render(request, 'pages/admin_painel.html', context)


# ============================================================
# 🔄 HTMX - OFICIAIS
# ============================================================
@login_required
def htmx_oficiais_lista(request):
    """Retorna a lista de oficiais para a tabela administrativa ou comparação."""
    
    user = request.user
    oficiais = Oficial.objects.all().order_by('posto', 'nome')
    
    # Se for comandante, filtrar apenas OBMs permitidas
    if user.is_comandante:
        obms_permitidas = user.get_obm_subordinadas()
        if obms_permitidas:
            q_filter = Q()
            for obm in obms_permitidas:
                q_filter |= Q(obm__icontains=obm)
            oficiais = oficiais.filter(q_filter)
    
    # Filtros da interface
    posto = request.GET.get('posto', '')
    quadro = request.GET.get('quadro', '')
    obm = request.GET.get('obm', '')
    busca = request.GET.get('busca', '')
    ativo = request.GET.get('ativo', '')
    
    if posto:
        oficiais = oficiais.filter(posto=posto)
    if quadro:
        oficiais = oficiais.filter(quadro=quadro)
    if obm:
        oficiais = oficiais.filter(obm__icontains=obm)
    if busca:
        oficiais = oficiais.filter(
            Q(nome__icontains=busca) | 
            Q(nome_guerra__icontains=busca) |
            Q(cpf__icontains=busca) |
            Q(rg__icontains=busca)
        )
    if ativo:
        oficiais = oficiais.filter(ativo=(ativo == 'true'))
    
    # Determinar qual template usar
    template = request.GET.get('template', 'tabela')
    
    if template == 'lista':
        return render(request, 'htmx/oficiais_lista.html', {'oficiais': oficiais})
    
    # Ordenação
    ordenar = request.GET.get('ordenar', 'posto')
    direcao = request.GET.get('direcao', 'asc')
    
    if direcao == 'desc' and not ordenar.startswith('-'):
        ordenar = f'-{ordenar}'
    elif direcao == 'asc' and ordenar.startswith('-'):
        ordenar = ordenar[1:]
    
    oficiais = oficiais.order_by(ordenar, 'nome')
    
    # Paginação
    por_pagina = int(request.GET.get('por_pagina', 25))
    pagina = request.GET.get('pagina', 1)
    
    paginator = Paginator(oficiais, por_pagina)
    page_obj = paginator.get_page(pagina)
    
    # Query string para paginação
    query_params = request.GET.copy()
    if 'pagina' in query_params:
        del query_params['pagina']
    query_string = query_params.urlencode()
    
    # Lista de OBMs disponíveis para filtro
    obms_disponiveis = Oficial.objects.values_list('obm', flat=True).distinct().order_by('obm')
    
    context = {
        'page_obj': page_obj,
        'filtros': {
            'busca': busca,
            'posto': posto,
            'quadro': quadro,
            'obm': obm,
            'ativo': ativo,
            'por_pagina': str(por_pagina),
        },
        'ordenacao': {
            'campo': ordenar.lstrip('-'),
            'direcao': direcao,
        },
        'query_string': query_string,
        'posto_choices': Oficial.POSTO_CHOICES,
        'quadro_choices': Oficial.QUADRO_CHOICES,
        'obms_disponiveis': obms_disponiveis,
        'user': user,
    }
    
    return render(request, 'htmx/oficiais_tabela.html', context)


@login_required
def htmx_oficiais_selecao(request):
    """Retorna lista de oficiais com checkboxes para seleção (página Comparar)."""
    
    user = request.user
    oficiais = Oficial.objects.filter(ativo=True)
    
    # Filtros
    posto = request.GET.get('posto', '')
    quadro = request.GET.get('quadro', '')
    obm = request.GET.get('obm', '')
    busca = request.GET.get('busca', '')
    
    if posto:
        oficiais = oficiais.filter(posto=posto)
    if quadro:
        oficiais = oficiais.filter(quadro=quadro)
    if obm:
        oficiais = oficiais.filter(obm__icontains=obm)
    if busca:
        oficiais = oficiais.filter(
            Q(nome__icontains=busca) | 
            Q(nome_guerra__icontains=busca)
        )
    
    # Filtro de OBM para comandante
    if user.role == 'comandante' and hasattr(user, 'get_obm_subordinadas'):
        obms_permitidas = user.get_obm_subordinadas()
        if obms_permitidas:
            q_filter = Q()
            for obm_perm in obms_permitidas:
                q_filter |= Q(obm__icontains=obm_perm)
            oficiais = oficiais.filter(q_filter)
    
    oficiais = oficiais.order_by('posto', 'nome')
    
    # Lista de OBMs disponíveis
    obms_disponiveis = Oficial.objects.filter(ativo=True).values_list('obm', flat=True).distinct().order_by('obm')
    
    context = {
        'oficiais': oficiais,
        'filtros': {
            'busca': busca,
            'posto': posto,
            'quadro': quadro,
            'obm': obm,
        },
        'posto_choices': Oficial.POSTO_CHOICES,
        'quadro_choices': Oficial.QUADRO_CHOICES,
        'obms_disponiveis': obms_disponiveis,
    }
    
    return render(request, 'htmx/oficiais_selecao.html', context)


@login_required
def htmx_oficiais_cards(request):
    """Retorna cards de oficiais para comparação com gráficos."""
    
    ids = request.GET.get('ids', '')
    
    if ids:
        ids_list = [int(id) for id in ids.split(',') if id.isdigit()]
        oficiais = Oficial.objects.filter(id__in=ids_list)
        
        # Para cada oficial, buscar dados para o card
        oficiais_data = []
        for oficial in oficiais:
            ultimas_missoes = oficial.designacoes.select_related('missao').filter(
                missao__status='EM_ANDAMENTO'
            ).order_by('-criado_em')[:5]
            
            oficiais_data.append({
                'oficial': oficial,
                'total_baixa': oficial.total_baixa,
                'total_media': oficial.total_media,
                'total_alta': oficial.total_alta,
                'carga_total': oficial.carga_total,
                'ultimas_missoes': ultimas_missoes,
            })
    else:
        oficiais_data = []
    
    return render(request, 'htmx/oficiais_cards.html', {'oficiais_data': oficiais_data})


@login_required
def htmx_oficial_card(request, pk):
    """Retorna o card de um oficial específico."""
    
    oficial = get_object_or_404(Oficial, pk=pk)
    
    # Últimas missões ativas
    ultimas_missoes = oficial.designacoes.select_related('missao').filter(
        missao__status='EM_ANDAMENTO'
    ).order_by('-criado_em')[:5]
    
    context = {
        'oficial': oficial,
        'total_baixa': oficial.total_baixa,
        'total_media': oficial.total_media,
        'total_alta': oficial.total_alta,
        'carga_total': oficial.carga_total,
        'ultimas_missoes': ultimas_missoes,
    }
    
    return render(request, 'htmx/oficial_card.html', context)


@login_required
@require_POST
def htmx_oficial_criar(request):
    """Cria um novo oficial via HTMX."""
    
    if not request.user.pode_gerenciar_oficiais:
        return HttpResponse('Sem permissão', status=403)
    
    try:
        oficial = Oficial.objects.create(
            cpf=request.POST.get('cpf', '').replace('.', '').replace('-', ''),
            rg=request.POST.get('rg', ''),
            nome=request.POST.get('nome', ''),
            nome_guerra=request.POST.get('nome_guerra', ''),
            posto=request.POST.get('posto', ''),
            quadro=request.POST.get('quadro', ''),
            obm=request.POST.get('obm', ''),
            funcao=request.POST.get('funcao', ''),
            email=request.POST.get('email', ''),
            telefone=request.POST.get('telefone', ''),
        )
        
        # Criar usuário automaticamente
        Usuario.objects.create_user(
            cpf=oficial.cpf,
            password='123456',
            oficial=oficial,
            role='oficial'
        )
        
        messages.success(request, f'Oficial {oficial} criado com sucesso!')
        
    except Exception as e:
        messages.error(request, f'Erro ao criar oficial: {str(e)}')
    
    return htmx_oficiais_lista(request)


@login_required
@require_POST
def htmx_oficial_editar(request, pk):
    """Edita um oficial via HTMX."""
    
    if not request.user.pode_gerenciar_oficiais:
        return HttpResponse('Sem permissão', status=403)
    
    oficial = get_object_or_404(Oficial, pk=pk)
    
    try:
        oficial.nome = request.POST.get('nome', oficial.nome)
        oficial.nome_guerra = request.POST.get('nome_guerra', oficial.nome_guerra)
        oficial.posto = request.POST.get('posto', oficial.posto)
        oficial.quadro = request.POST.get('quadro', oficial.quadro)
        oficial.obm = request.POST.get('obm', oficial.obm)
        oficial.funcao = request.POST.get('funcao', oficial.funcao)
        oficial.email = request.POST.get('email', oficial.email)
        oficial.telefone = request.POST.get('telefone', oficial.telefone)
        oficial.save()
        
        messages.success(request, f'Oficial {oficial} atualizado!')
        
    except Exception as e:
        messages.error(request, f'Erro ao atualizar: {str(e)}')
    
    return htmx_oficiais_lista(request)


@login_required
def htmx_oficial_dados(request, pk):
    """Retorna dados de um oficial em JSON para edição."""
    
    oficial = get_object_or_404(Oficial, pk=pk)
    
    return JsonResponse({
        'id': oficial.id,
        'cpf': oficial.cpf,
        'rg': oficial.rg,
        'nome': oficial.nome,
        'nome_guerra': oficial.nome_guerra,
        'posto': oficial.posto,
        'quadro': oficial.quadro,
        'obm': oficial.obm,
        'funcao': oficial.funcao,
        'email': oficial.email,
        'telefone': oficial.telefone,
        'ativo': oficial.ativo,
    })


@login_required
@require_POST
def htmx_oficial_excluir(request, pk):
    """Exclui um oficial via HTMX."""
    
    if not request.user.pode_gerenciar_oficiais:
        return HttpResponse('Sem permissão', status=403)
    
    oficial = get_object_or_404(Oficial, pk=pk)
    nome = str(oficial)
    
    try:
        oficial.delete()
        messages.success(request, f'Oficial {nome} excluído!')
    except Exception as e:
        messages.error(request, f'Erro ao excluir: {str(e)}')
    
    return htmx_oficiais_lista(request)


# ============================================================
# 🔄 HTMX - MISSÕES
# ============================================================
@login_required
def htmx_missoes_lista(request):
    """Retorna a lista de missões filtrada (cards para página de Missões)."""
    
    missoes = Missao.objects.all().order_by('-data_inicio')
    
    # Filtros
    tipo = request.GET.get('tipo', '')
    status = request.GET.get('status', '')
    local = request.GET.get('local', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    
    if tipo:
        missoes = missoes.filter(tipo=tipo)
    if status:
        missoes = missoes.filter(status=status)
    if local:
        missoes = missoes.filter(local__icontains=local)
    if data_inicio:
        missoes = missoes.filter(data_inicio__gte=data_inicio)
    if data_fim:
        missoes = missoes.filter(data_fim__lte=data_fim)
    
    return render(request, 'htmx/missoes_lista.html', {'missoes': missoes})


@login_required
def htmx_missoes_tabela(request):
    """Retorna tabela de missões com paginação e filtros (para Admin)."""
    
    missoes = Missao.objects.all()
    
    # ============================================================
    # FILTROS
    # ============================================================
    busca = request.GET.get('busca', '').strip()
    tipo = request.GET.get('tipo', '')
    status = request.GET.get('status', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    
    if busca:
        missoes = missoes.filter(
            Q(nome__icontains=busca) |
            Q(local__icontains=busca) |
            Q(documento_referencia__icontains=busca)
        )
    
    if tipo:
        missoes = missoes.filter(tipo=tipo)
    if status:
        missoes = missoes.filter(status=status)
    if data_inicio:
        missoes = missoes.filter(data_inicio__gte=data_inicio)
    if data_fim:
        missoes = missoes.filter(data_fim__lte=data_fim)
    
    # ============================================================
    # ORDENAÇÃO
    # ============================================================
    ordenar = request.GET.get('ordenar', '-data_inicio')
    direcao = request.GET.get('direcao', 'desc')
    
    if direcao == 'desc' and not ordenar.startswith('-'):
        ordenar = f'-{ordenar}'
    elif direcao == 'asc' and ordenar.startswith('-'):
        ordenar = ordenar[1:]
    
    missoes = missoes.order_by(ordenar)
    
    # ============================================================
    # PAGINAÇÃO
    # ============================================================
    por_pagina = int(request.GET.get('por_pagina', 25))
    pagina = request.GET.get('pagina', 1)
    
    paginator = Paginator(missoes, por_pagina)
    page_obj = paginator.get_page(pagina)
    
    # Query string para paginação
    query_params = request.GET.copy()
    if 'pagina' in query_params:
        del query_params['pagina']
    query_string = query_params.urlencode()
    
    context = {
        'page_obj': page_obj,
        'filtros': {
            'busca': busca,
            'tipo': tipo,
            'status': status,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'por_pagina': str(por_pagina),
        },
        'ordenacao': {
            'campo': ordenar.lstrip('-'),
            'direcao': direcao,
        },
        'query_string': query_string,
        'tipo_choices': Missao.TIPO_CHOICES,
        'status_choices': Missao.STATUS_CHOICES,
        'user': request.user,
    }
    
    return render(request, 'htmx/missoes_tabela.html', context)


@login_required
def htmx_missao_organograma(request, pk):
    """Retorna o organograma de uma missão."""
    
    missao = get_object_or_404(Missao, pk=pk)
    
    designacoes = missao.designacoes.select_related('oficial').all()
    
    # Separar por hierarquia
    superiores = designacoes.filter(
        funcao_na_missao__in=['COMANDANTE', 'PRESIDENTE', 'COORDENADOR', 'ENCARREGADO']
    )
    subordinados = designacoes.exclude(
        funcao_na_missao__in=['COMANDANTE', 'PRESIDENTE', 'COORDENADOR', 'ENCARREGADO']
    )
    
    context = {
        'missao': missao,
        'superiores': superiores,
        'subordinados': subordinados,
    }
    
    return render(request, 'htmx/missao_organograma.html', context)


@login_required
@require_POST
def htmx_missao_criar(request):
    """Cria uma nova missão via HTMX."""
    
    if not request.user.pode_gerenciar_missoes:
        return HttpResponse('Sem permissão', status=403)
    
    try:
        Missao.objects.create(
            tipo=request.POST.get('tipo', ''),
            nome=request.POST.get('nome', ''),
            descricao=request.POST.get('descricao', ''),
            local=request.POST.get('local', ''),
            data_inicio=request.POST.get('data_inicio') or None,
            data_fim=request.POST.get('data_fim') or None,
            status=request.POST.get('status', 'PLANEJADA'),
            documento_referencia=request.POST.get('documento_referencia', ''),
        )
        messages.success(request, 'Missão criada com sucesso!')
        
    except Exception as e:
        messages.error(request, f'Erro ao criar missão: {str(e)}')
    
    return htmx_missoes_lista(request)


@login_required
@require_POST
def htmx_missao_editar(request, pk):
    """Edita uma missão via HTMX."""
    
    if not request.user.pode_gerenciar_missoes:
        return HttpResponse('Sem permissão', status=403)
    
    missao = get_object_or_404(Missao, pk=pk)
    
    try:
        missao.tipo = request.POST.get('tipo', missao.tipo)
        missao.nome = request.POST.get('nome', missao.nome)
        missao.descricao = request.POST.get('descricao', missao.descricao)
        missao.local = request.POST.get('local', missao.local)
        missao.status = request.POST.get('status', missao.status)
        missao.documento_referencia = request.POST.get('documento_referencia', missao.documento_referencia)
        
        data_inicio = request.POST.get('data_inicio')
        data_fim = request.POST.get('data_fim')
        if data_inicio:
            missao.data_inicio = data_inicio
        if data_fim:
            missao.data_fim = data_fim
        
        missao.save()
        messages.success(request, 'Missão atualizada!')
        
    except Exception as e:
        messages.error(request, f'Erro ao atualizar: {str(e)}')
    
    return htmx_missoes_lista(request)


@login_required
def htmx_missao_dados(request, pk):
    """Retorna dados de uma missão em JSON para edição."""
    
    missao = get_object_or_404(Missao, pk=pk)
    
    return JsonResponse({
        'id': missao.id,
        'nome': missao.nome,
        'tipo': missao.tipo,
        'status': missao.status,
        'descricao': missao.descricao,
        'local': missao.local,
        'data_inicio': missao.data_inicio.strftime('%Y-%m-%d') if missao.data_inicio else '',
        'data_fim': missao.data_fim.strftime('%Y-%m-%d') if missao.data_fim else '',
        'documento_referencia': missao.documento_referencia,
    })


@login_required
@require_POST
def htmx_missao_excluir(request, pk):
    """Exclui uma missão via HTMX."""
    
    if not request.user.pode_gerenciar_missoes:
        return HttpResponse('Sem permissão', status=403)
    
    missao = get_object_or_404(Missao, pk=pk)
    nome = missao.nome
    
    try:
        missao.delete()
        messages.success(request, f'Missão "{nome}" excluída!')
    except Exception as e:
        messages.error(request, f'Erro ao excluir: {str(e)}')
    
    return htmx_missoes_lista(request)


# ============================================================
# 🔄 HTMX - DESIGNAÇÕES
# ============================================================
@login_required
def htmx_designacoes_lista(request):
    """Retorna a lista de designações com paginação e filtros (para Admin)."""
    
    designacoes = Designacao.objects.select_related('missao', 'oficial').all()
    
    # ============================================================
    # FILTROS
    # ============================================================
    busca = request.GET.get('busca', '').strip()
    missao_id = request.GET.get('missao_id', '')
    funcao = request.GET.get('funcao', '')
    complexidade = request.GET.get('complexidade', '')
    
    if busca:
        designacoes = designacoes.filter(
            Q(oficial__nome__icontains=busca) |
            Q(oficial__nome_guerra__icontains=busca) |
            Q(missao__nome__icontains=busca)
        )
    
    if missao_id:
        designacoes = designacoes.filter(missao_id=missao_id)
    
    if funcao:
        designacoes = designacoes.filter(funcao_na_missao=funcao)
    
    if complexidade:
        designacoes = designacoes.filter(complexidade=complexidade)
    
    # ============================================================
    # ORDENAÇÃO
    # ============================================================
    ordenar = request.GET.get('ordenar', '-criado_em')
    direcao = request.GET.get('direcao', 'desc')
    
    if direcao == 'desc' and not ordenar.startswith('-'):
        ordenar = f'-{ordenar}'
    elif direcao == 'asc' and ordenar.startswith('-'):
        ordenar = ordenar[1:]
    
    designacoes = designacoes.order_by(ordenar)
    
    # ============================================================
    # PAGINAÇÃO
    # ============================================================
    por_pagina = int(request.GET.get('por_pagina', 25))
    pagina = request.GET.get('pagina', 1)
    
    paginator = Paginator(designacoes, por_pagina)
    page_obj = paginator.get_page(pagina)
    
    # Query string para paginação
    query_params = request.GET.copy()
    if 'pagina' in query_params:
        del query_params['pagina']
    query_string = query_params.urlencode()
    
    context = {
        'page_obj': page_obj,
        'filtros': {
            'busca': busca,
            'missao_id': missao_id,
            'funcao': funcao,
            'complexidade': complexidade,
            'por_pagina': str(por_pagina),
        },
        'ordenacao': {
            'campo': ordenar.lstrip('-'),
            'direcao': direcao,
        },
        'query_string': query_string,
        'funcao_choices': Designacao.FUNCAO_CHOICES,
        'complexidade_choices': Designacao.COMPLEXIDADE_CHOICES,
        'missoes_disponiveis': Missao.objects.filter(status__in=['PLANEJADA', 'EM_ANDAMENTO']).order_by('nome'),
        'oficiais_disponiveis': Oficial.objects.filter(ativo=True).order_by('posto', 'nome'),
        'user': request.user,
    }
    
    return render(request, 'htmx/designacoes_tabela.html', context)


@login_required
@require_POST
def htmx_designacao_criar(request):
    """Cria uma nova designação via HTMX."""
    
    if not request.user.pode_gerenciar_designacoes:
        return HttpResponse('Sem permissão', status=403)
    
    try:
        missao_id = request.POST.get('missao_id')
        oficial_id = request.POST.get('oficial_id')
        
        Designacao.objects.create(
            missao_id=missao_id,
            oficial_id=oficial_id,
            funcao_na_missao=request.POST.get('funcao_na_missao', 'MEMBRO'),
            complexidade=request.POST.get('complexidade', 'MEDIA'),
            observacoes=request.POST.get('observacoes', ''),
        )
        messages.success(request, 'Designação criada!')
        
    except Exception as e:
        messages.error(request, f'Erro ao criar designação: {str(e)}')
    
    return htmx_designacoes_lista(request)


@login_required
@require_POST
def htmx_designacao_editar(request, pk):
    """Edita uma designação via HTMX."""
    
    if not request.user.pode_gerenciar_designacoes:
        return HttpResponse('Sem permissão', status=403)
    
    designacao = get_object_or_404(Designacao, pk=pk)
    
    try:
        designacao.funcao_na_missao = request.POST.get('funcao_na_missao', designacao.funcao_na_missao)
        designacao.complexidade = request.POST.get('complexidade', designacao.complexidade)
        designacao.observacoes = request.POST.get('observacoes', designacao.observacoes)
        designacao.save()
        messages.success(request, 'Designação atualizada!')
        
    except Exception as e:
        messages.error(request, f'Erro ao atualizar: {str(e)}')
    
    return htmx_designacoes_lista(request)


@login_required
def htmx_designacao_dados(request, pk):
    """Retorna dados de uma designação em JSON para edição."""
    
    designacao = get_object_or_404(Designacao, pk=pk)
    
    return JsonResponse({
        'id': designacao.id,
        'missao_id': designacao.missao_id,
        'oficial_id': designacao.oficial_id,
        'funcao_na_missao': designacao.funcao_na_missao,
        'complexidade': designacao.complexidade,
        'observacoes': designacao.observacoes,
    })


@login_required
@require_POST
def htmx_designacao_excluir(request, pk):
    """Exclui uma designação via HTMX."""
    
    if not request.user.pode_gerenciar_designacoes:
        return HttpResponse('Sem permissão', status=403)
    
    designacao = get_object_or_404(Designacao, pk=pk)
    
    try:
        designacao.delete()
        messages.success(request, 'Designação excluída!')
    except Exception as e:
        messages.error(request, f'Erro ao excluir: {str(e)}')
    
    return htmx_designacoes_lista(request)


# ============================================================
# 🔄 HTMX - UNIDADES
# ============================================================
@login_required
def htmx_unidades_lista(request):
    """Retorna a lista de unidades com paginação e filtros."""
    
    unidades = Unidade.objects.all()
    
    # Filtros
    busca = request.GET.get('busca', '').strip()
    tipo = request.GET.get('tipo', '')
    
    if busca:
        unidades = unidades.filter(
            Q(nome__icontains=busca) |
            Q(sigla__icontains=busca)
        )
    
    if tipo:
        unidades = unidades.filter(tipo=tipo)
    
    # Ordenação
    ordenar = request.GET.get('ordenar', 'nome')
    direcao = request.GET.get('direcao', 'asc')
    
    if direcao == 'desc' and not ordenar.startswith('-'):
        ordenar = f'-{ordenar}'
    elif direcao == 'asc' and ordenar.startswith('-'):
        ordenar = ordenar[1:]
    
    unidades = unidades.order_by(ordenar)
    
    # Paginação
    por_pagina = int(request.GET.get('por_pagina', 25))
    pagina = request.GET.get('pagina', 1)
    
    paginator = Paginator(unidades, por_pagina)
    page_obj = paginator.get_page(pagina)
    
    # Query string
    query_params = request.GET.copy()
    if 'pagina' in query_params:
        del query_params['pagina']
    query_string = query_params.urlencode()
    
    context = {
        'page_obj': page_obj,
        'filtros': {
            'busca': busca,
            'tipo': tipo,
            'por_pagina': str(por_pagina),
        },
        'ordenacao': {
            'campo': ordenar.lstrip('-'),
            'direcao': direcao,
        },
        'query_string': query_string,
        'tipo_choices': Unidade.TIPO_CHOICES,
        'unidades_disponiveis': Unidade.objects.all().order_by('nome'),
        'user': request.user,
    }
    
    return render(request, 'htmx/unidades_tabela.html', context)


@login_required
@require_POST
def htmx_unidade_criar(request):
    """Cria uma nova unidade via HTMX."""
    
    if not request.user.pode_gerenciar_unidades:
        return HttpResponse('Sem permissão', status=403)
    
    try:
        comando_superior_id = request.POST.get('comando_superior_id')
        
        Unidade.objects.create(
            nome=request.POST.get('nome', ''),
            sigla=request.POST.get('sigla', ''),
            tipo=request.POST.get('tipo', ''),
            comando_superior_id=comando_superior_id if comando_superior_id else None,
        )
        messages.success(request, 'Unidade criada!')
        
    except Exception as e:
        messages.error(request, f'Erro ao criar unidade: {str(e)}')
    
    return htmx_unidades_lista(request)


@login_required
@require_POST
def htmx_unidade_editar(request, pk):
    """Edita uma unidade via HTMX."""
    
    if not request.user.pode_gerenciar_unidades:
        return HttpResponse('Sem permissão', status=403)
    
    unidade = get_object_or_404(Unidade, pk=pk)
    
    try:
        unidade.nome = request.POST.get('nome', unidade.nome)
        unidade.sigla = request.POST.get('sigla', unidade.sigla)
        unidade.tipo = request.POST.get('tipo', unidade.tipo)
        
        comando_superior_id = request.POST.get('comando_superior_id')
        unidade.comando_superior_id = comando_superior_id if comando_superior_id else None
        
        unidade.save()
        messages.success(request, 'Unidade atualizada!')
        
    except Exception as e:
        messages.error(request, f'Erro ao atualizar: {str(e)}')
    
    return htmx_unidades_lista(request)


@login_required
@require_POST
def htmx_unidade_excluir(request, pk):
    """Exclui uma unidade via HTMX."""
    
    if not request.user.pode_gerenciar_unidades:
        return HttpResponse('Sem permissão', status=403)
    
    unidade = get_object_or_404(Unidade, pk=pk)
    nome = unidade.nome
    
    try:
        unidade.delete()
        messages.success(request, f'Unidade "{nome}" excluída!')
    except Exception as e:
        messages.error(request, f'Erro ao excluir: {str(e)}')
    
    return htmx_unidades_lista(request)


# ============================================================
# 🔄 HTMX - USUÁRIOS
# ============================================================
@login_required
def htmx_usuarios_lista(request):
    """Retorna a lista de usuários com paginação e filtros."""
    
    if not request.user.pode_gerenciar_usuarios:
        return HttpResponse('Sem permissão', status=403)
    
    usuarios = Usuario.objects.select_related('oficial').all()
    
    # Filtros
    busca = request.GET.get('busca', '').strip()
    role = request.GET.get('role', '')
    ativo = request.GET.get('ativo', '')
    
    if busca:
        usuarios = usuarios.filter(
            Q(cpf__icontains=busca) |
            Q(oficial__nome__icontains=busca) |
            Q(oficial__nome_guerra__icontains=busca)
        )
    
    if role:
        usuarios = usuarios.filter(role=role)
    
    if ativo:
        usuarios = usuarios.filter(is_active=(ativo == 'true'))
    
    # Ordenação
    ordenar = request.GET.get('ordenar', 'cpf')
    direcao = request.GET.get('direcao', 'asc')
    
    if direcao == 'desc' and not ordenar.startswith('-'):
        ordenar = f'-{ordenar}'
    elif direcao == 'asc' and ordenar.startswith('-'):
        ordenar = ordenar[1:]
    
    usuarios = usuarios.order_by(ordenar)
    
    # Paginação
    por_pagina = int(request.GET.get('por_pagina', 25))
    pagina = request.GET.get('pagina', 1)
    
    paginator = Paginator(usuarios, por_pagina)
    page_obj = paginator.get_page(pagina)
    
    # Query string
    query_params = request.GET.copy()
    if 'pagina' in query_params:
        del query_params['pagina']
    query_string = query_params.urlencode()
    
    # Oficiais sem usuário (para vincular)
    oficiais_disponiveis = Oficial.objects.filter(usuario__isnull=True).order_by('posto', 'nome')
    
    context = {
        'page_obj': page_obj,
        'filtros': {
            'busca': busca,
            'role': role,
            'ativo': ativo,
            'por_pagina': str(por_pagina),
        },
        'ordenacao': {
            'campo': ordenar.lstrip('-'),
            'direcao': direcao,
        },
        'query_string': query_string,
        'role_choices': Usuario.ROLE_CHOICES,
        'oficiais_disponiveis': oficiais_disponiveis,
        'user': request.user,
    }
    
    return render(request, 'htmx/usuarios_tabela.html', context)


@login_required
@require_POST
def htmx_usuario_criar(request):
    """Cria um novo usuário via HTMX."""
    
    if not request.user.is_admin:
        return HttpResponse('Sem permissão', status=403)
    
    try:
        cpf = request.POST.get('cpf', '').replace('.', '').replace('-', '')
        oficial_id = request.POST.get('oficial_id')
        
        Usuario.objects.create_user(
            cpf=cpf,
            password='123456',
            role=request.POST.get('role', 'oficial'),
            oficial_id=oficial_id if oficial_id else None,
        )
        messages.success(request, f'Usuário {cpf} criado com senha padrão 123456!')
        
    except Exception as e:
        messages.error(request, f'Erro ao criar usuário: {str(e)}')
    
    return htmx_usuarios_lista(request)


@login_required
@require_POST
@login_required
@require_POST
def htmx_usuario_editar(request, pk):
    """Edita um usuário via HTMX."""
    
    if not request.user.is_admin:
        return HttpResponse('Sem permissão', status=403)
    
    usuario = get_object_or_404(Usuario, pk=pk)
    
    try:
        usuario.role = request.POST.get('role', usuario.role)
        usuario.is_active = request.POST.get('is_active') == 'on'
        usuario.save()
        messages.success(request, 'Usuário atualizado!')
        
    except Exception as e:
        messages.error(request, f'Erro ao atualizar: {str(e)}')
    
    return htmx_usuarios_lista(request)


@login_required
@require_POST
def htmx_usuario_excluir(request, pk):
    """Exclui um usuário via HTMX."""
    
    if not request.user.is_admin:
        return HttpResponse('Sem permissão', status=403)
    
    usuario = get_object_or_404(Usuario, pk=pk)
    
    if usuario == request.user:
        messages.error(request, 'Você não pode excluir seu próprio usuário!')
        return htmx_usuarios_lista(request)
    
    try:
        usuario.delete()
        messages.success(request, 'Usuário excluído!')
    except Exception as e:
        messages.error(request, f'Erro ao excluir: {str(e)}')
    
    return htmx_usuarios_lista(request)


@login_required
@require_POST
def htmx_usuario_reset_senha(request, pk):
    """Reseta a senha de um usuário para 123456."""
    
    if not request.user.is_admin:
        return HttpResponse('Sem permissão', status=403)
    
    usuario = get_object_or_404(Usuario, pk=pk)
    
    try:
        usuario.set_password('123456')
        usuario.save()
        messages.success(request, f'Senha do usuário {usuario.cpf} redefinida para 123456!')
    except Exception as e:
        messages.error(request, f'Erro ao resetar senha: {str(e)}')
    
    return htmx_usuarios_lista(request)


# ============================================================
# 🔄 HTMX - SOLICITAÇÕES
# ============================================================
@login_required
@require_POST
def htmx_solicitacao_missao_criar(request):
    """Cria uma solicitação de inclusão de missão."""
    
    if not request.user.oficial:
        messages.error(request, 'Usuário não vinculado a um oficial.')
        return HttpResponse('<div class="alert alert-danger">Usuário não vinculado a um oficial.</div>')
    
    try:
        # Processar datas
        data_inicio = request.POST.get('data_inicio')
        data_fim = request.POST.get('data_fim') or None
        status_missao = request.POST.get('status_missao', 'EM_ANDAMENTO')
        
        # Se status for CONCLUIDA, data_fim é obrigatória
        if status_missao == 'CONCLUIDA' and not data_fim:
            return HttpResponse('<div class="alert alert-danger">Data de término é obrigatória para missões concluídas.</div>')
        
        SolicitacaoMissao.objects.create(
            solicitante=request.user.oficial,
            nome_missao=request.POST.get('nome_missao', ''),
            tipo_missao=request.POST.get('tipo_missao', ''),
            status_missao=status_missao,
            local=request.POST.get('local', ''),
            data_inicio=data_inicio,
            data_fim=data_fim,
            documento_sei=request.POST.get('documento_sei', ''),
        )
        
        return HttpResponse('<div class="alert alert-success"><i data-lucide="check-circle"></i> Solicitação de missão enviada com sucesso! Aguarde avaliação.</div><script>lucide.createIcons();</script>')
        
    except Exception as e:
        return HttpResponse(f'<div class="alert alert-danger">Erro ao criar solicitação: {str(e)}</div>')


@login_required
@require_POST
def htmx_solicitacao_designacao_criar(request):
    """Cria uma solicitação de inclusão de designação."""
    
    if not request.user.oficial:
        messages.error(request, 'Usuário não vinculado a um oficial.')
        return HttpResponse('<div class="alert alert-danger">Usuário não vinculado a um oficial.</div>')
    
    try:
        missao_id = request.POST.get('missao_id')
        if not missao_id:
            return HttpResponse('<div class="alert alert-danger">Selecione uma missão.</div>')
        
        missao = Missao.objects.get(id=missao_id)
        
        # Verificar se já existe designação para este oficial nesta missão
        if Designacao.objects.filter(oficial=request.user.oficial, missao=missao).exists():
            return HttpResponse('<div class="alert alert-warning">Você já está designado para esta missão.</div>')
        
        # Verificar se já existe solicitação pendente
        if SolicitacaoDesignacao.objects.filter(
            solicitante=request.user.oficial, 
            missao=missao, 
            status='PENDENTE'
        ).exists():
            return HttpResponse('<div class="alert alert-warning">Já existe uma solicitação pendente para esta missão.</div>')
        
        SolicitacaoDesignacao.objects.create(
            solicitante=request.user.oficial,
            missao=missao,
            funcao_na_missao=request.POST.get('funcao_na_missao', ''),
            documento_sei=request.POST.get('documento_sei', ''),
        )
        
        return HttpResponse('<div class="alert alert-success"><i data-lucide="check-circle"></i> Solicitação de designação enviada com sucesso! Aguarde avaliação.</div><script>lucide.createIcons();</script>')
        
    except Missao.DoesNotExist:
        return HttpResponse('<div class="alert alert-danger">Missão não encontrada.</div>')
    except Exception as e:
        return HttpResponse(f'<div class="alert alert-danger">Erro ao criar solicitação: {str(e)}</div>')


@login_required
def htmx_solicitacoes_lista(request):
    """Lista solicitações de missão e designação com paginação e filtros."""
    
    if not request.user.pode_gerenciar_solicitacoes:
        return HttpResponse('Sem permissão', status=403)
    
    # Buscar solicitações de ambos os tipos
    tipo_solicitacao = request.GET.get('tipo_solicitacao', 'todas')
    
    # Filtros comuns
    busca = request.GET.get('busca', '').strip()
    status = request.GET.get('status', '')
    
    # Listas separadas
    solicitacoes_missao = []
    solicitacoes_designacao = []
    
    if tipo_solicitacao in ['todas', 'missao']:
        qs_missao = SolicitacaoMissao.objects.select_related('solicitante', 'avaliado_por').all()
        if busca:
            qs_missao = qs_missao.filter(
                Q(solicitante__nome__icontains=busca) |
                Q(solicitante__nome_guerra__icontains=busca) |
                Q(nome_missao__icontains=busca)
            )
        if status:
            qs_missao = qs_missao.filter(status=status)
        solicitacoes_missao = list(qs_missao.order_by('-criado_em'))
    
    if tipo_solicitacao in ['todas', 'designacao']:
        qs_designacao = SolicitacaoDesignacao.objects.select_related('solicitante', 'avaliado_por', 'missao').all()
        if busca:
            qs_designacao = qs_designacao.filter(
                Q(solicitante__nome__icontains=busca) |
                Q(solicitante__nome_guerra__icontains=busca) |
                Q(missao__nome__icontains=busca)
            )
        if status:
            qs_designacao = qs_designacao.filter(status=status)
        solicitacoes_designacao = list(qs_designacao.order_by('-criado_em'))
    
    # Contadores
    pendentes_missao = SolicitacaoMissao.objects.filter(status='PENDENTE').count()
    pendentes_designacao = SolicitacaoDesignacao.objects.filter(status='PENDENTE').count()
    
    context = {
        'solicitacoes_missao': solicitacoes_missao,
        'solicitacoes_designacao': solicitacoes_designacao,
        'pendentes_missao': pendentes_missao,
        'pendentes_designacao': pendentes_designacao,
        'filtros': {
            'busca': busca,
            'status': status,
            'tipo_solicitacao': tipo_solicitacao,
        },
        'status_choices': SolicitacaoDesignacao.STATUS_CHOICES,
        'complexidade_choices': Designacao.COMPLEXIDADE_CHOICES,
        'user': request.user,
    }
    
    return render(request, 'htmx/solicitacoes_lista.html', context)


@login_required
@require_POST
def htmx_solicitacao_missao_avaliar(request, pk):
    """Avalia uma solicitação de missão (aprovar/recusar)."""
    
    if not request.user.pode_gerenciar_solicitacoes:
        return HttpResponse('Sem permissão', status=403)
    
    solicitacao = get_object_or_404(SolicitacaoMissao, pk=pk)
    acao = request.POST.get('acao')  # 'aprovar' ou 'recusar'
    
    try:
        solicitacao.avaliado_por = request.user
        solicitacao.data_avaliacao = timezone.now()
        solicitacao.observacao_avaliador = request.POST.get('observacao', '')
        
        if acao == 'aprovar':
            # Criar a missão automaticamente
            missao = Missao.objects.create(
                nome=solicitacao.nome_missao,
                tipo=solicitacao.tipo_missao,
                status=solicitacao.status_missao,
                local=dict(SolicitacaoMissao.LOCAL_CHOICES).get(solicitacao.local, solicitacao.local),
                data_inicio=solicitacao.data_inicio,
                data_fim=solicitacao.data_fim,
                documento_referencia=solicitacao.documento_sei,
            )
            solicitacao.status = 'APROVADA'
            solicitacao.missao_criada = missao
            messages.success(request, f'Solicitação aprovada! Missão "{missao.nome}" criada com sucesso.')
        else:
            solicitacao.status = 'RECUSADA'
            messages.info(request, 'Solicitação recusada.')
        
        solicitacao.save()
        
    except Exception as e:
        messages.error(request, f'Erro ao avaliar: {str(e)}')
    
    return htmx_solicitacoes_lista(request)


@login_required
@require_POST
def htmx_solicitacao_designacao_avaliar(request, pk):
    """Avalia uma solicitação de designação (aprovar/recusar)."""
    
    if not request.user.pode_gerenciar_solicitacoes:
        return HttpResponse('Sem permissão', status=403)
    
    solicitacao = get_object_or_404(SolicitacaoDesignacao, pk=pk)
    acao = request.POST.get('acao')  # 'aprovar' ou 'recusar'
    
    try:
        solicitacao.avaliado_por = request.user
        solicitacao.data_avaliacao = timezone.now()
        solicitacao.observacao_avaliador = request.POST.get('observacao', '')
        
        if acao == 'aprovar':
            # Complexidade é definida pelo BM/3 na aprovação
            complexidade = request.POST.get('complexidade', 'MEDIA')
            solicitacao.complexidade = complexidade
            
            # Criar a designação automaticamente
            designacao = Designacao.objects.create(
                oficial=solicitacao.solicitante,
                missao=solicitacao.missao,
                funcao_na_missao=solicitacao.funcao_na_missao,
                complexidade=complexidade,
                observacoes=f'Criado via solicitação. SEI: {solicitacao.documento_sei}',
            )
            solicitacao.status = 'APROVADA'
            solicitacao.designacao_criada = designacao
            messages.success(request, f'Solicitação aprovada! {solicitacao.solicitante} designado para "{solicitacao.missao.nome}".')
        else:
            solicitacao.status = 'RECUSADA'
            messages.info(request, 'Solicitação recusada.')
        
        solicitacao.save()
        
    except Exception as e:
        messages.error(request, f'Erro ao avaliar: {str(e)}')
    
    return htmx_solicitacoes_lista(request)


@login_required
def htmx_solicitacao_missao_dados(request, pk):
    """Retorna os dados de uma solicitação de missão para edição."""
    
    if not request.user.pode_gerenciar_solicitacoes:
        return HttpResponse('Sem permissão', status=403)
    
    solicitacao = get_object_or_404(SolicitacaoMissao, pk=pk)
    
    context = {
        'solicitacao': solicitacao,
        'tipo_choices': Missao.TIPO_CHOICES,
        'status_missao_choices': Missao.STATUS_CHOICES,
        'local_choices': SolicitacaoMissao.LOCAL_CHOICES,
    }
    
    return render(request, 'htmx/solicitacao_missao_form.html', context)


@login_required
@require_POST
def htmx_solicitacao_missao_editar(request, pk):
    """Edita uma solicitação de missão."""
    
    if not request.user.pode_gerenciar_solicitacoes:
        return HttpResponse('Sem permissão', status=403)
    
    solicitacao = get_object_or_404(SolicitacaoMissao, pk=pk)
    
    try:
        solicitacao.nome_missao = request.POST.get('nome_missao', solicitacao.nome_missao)
        solicitacao.tipo_missao = request.POST.get('tipo_missao', solicitacao.tipo_missao)
        solicitacao.status_missao = request.POST.get('status_missao', solicitacao.status_missao)
        solicitacao.local = request.POST.get('local', solicitacao.local)
        solicitacao.data_inicio = request.POST.get('data_inicio', solicitacao.data_inicio)
        data_fim = request.POST.get('data_fim')
        solicitacao.data_fim = data_fim if data_fim else None
        solicitacao.documento_sei = request.POST.get('documento_sei', solicitacao.documento_sei)
        solicitacao.save()
        
        return HttpResponse('<div class="alert alert-success"><i data-lucide="check-circle"></i> Solicitação atualizada com sucesso!</div><script>lucide.createIcons(); setTimeout(() => location.reload(), 1000);</script>')
        
    except Exception as e:
        return HttpResponse(f'<div class="alert alert-danger">Erro ao atualizar: {str(e)}</div>')


@login_required
def htmx_solicitacao_designacao_dados(request, pk):
    """Retorna os dados de uma solicitação de designação para edição."""
    
    if not request.user.pode_gerenciar_solicitacoes:
        return HttpResponse('Sem permissão', status=403)
    
    solicitacao = get_object_or_404(SolicitacaoDesignacao, pk=pk)
    missoes = Missao.objects.filter(status__in=['PLANEJADA', 'EM_ANDAMENTO']).order_by('nome')
    
    context = {
        'solicitacao': solicitacao,
        'missoes': missoes,
    }
    
    return render(request, 'htmx/solicitacao_designacao_form.html', context)


@login_required
@require_POST
def htmx_solicitacao_designacao_editar(request, pk):
    """Edita uma solicitação de designação."""
    
    if not request.user.pode_gerenciar_solicitacoes:
        return HttpResponse('Sem permissão', status=403)
    
    solicitacao = get_object_or_404(SolicitacaoDesignacao, pk=pk)
    
    try:
        missao_id = request.POST.get('missao_id')
        if missao_id:
            solicitacao.missao = Missao.objects.get(id=missao_id)
        solicitacao.funcao_na_missao = request.POST.get('funcao_na_missao', solicitacao.funcao_na_missao)
        solicitacao.documento_sei = request.POST.get('documento_sei', solicitacao.documento_sei)
        solicitacao.save()
        
        return HttpResponse('<div class="alert alert-success"><i data-lucide="check-circle"></i> Solicitação atualizada com sucesso!</div><script>lucide.createIcons(); setTimeout(() => location.reload(), 1000);</script>')
        
    except Exception as e:
        return HttpResponse(f'<div class="alert alert-danger">Erro ao atualizar: {str(e)}</div>')


@login_required
def minhas_solicitacoes(request):
    """Página com histórico de solicitações do oficial logado."""
    
    if not request.user.oficial:
        messages.error(request, 'Você não possui um oficial vinculado ao seu usuário.')
        return redirect('missoes_dashboard')
    
    oficial = request.user.oficial
    
    # Filtros
    tipo = request.GET.get('tipo', 'todas')
    status = request.GET.get('status', '')
    
    # Buscar solicitações
    solicitacoes_missao = SolicitacaoMissao.objects.filter(solicitante=oficial)
    solicitacoes_designacao = SolicitacaoDesignacao.objects.filter(solicitante=oficial).select_related('missao')
    
    if status:
        solicitacoes_missao = solicitacoes_missao.filter(status=status)
        solicitacoes_designacao = solicitacoes_designacao.filter(status=status)
    
    solicitacoes_missao = solicitacoes_missao.order_by('-criado_em')
    solicitacoes_designacao = solicitacoes_designacao.order_by('-criado_em')
    
    # Contadores
    total_missao = SolicitacaoMissao.objects.filter(solicitante=oficial).count()
    total_designacao = SolicitacaoDesignacao.objects.filter(solicitante=oficial).count()
    pendentes = (
        SolicitacaoMissao.objects.filter(solicitante=oficial, status='PENDENTE').count() +
        SolicitacaoDesignacao.objects.filter(solicitante=oficial, status='PENDENTE').count()
    )
    
    context = {
        'oficial': oficial,
        'solicitacoes_missao': solicitacoes_missao if tipo in ['todas', 'missao'] else [],
        'solicitacoes_designacao': solicitacoes_designacao if tipo in ['todas', 'designacao'] else [],
        'total_missao': total_missao,
        'total_designacao': total_designacao,
        'pendentes': pendentes,
        'filtros': {
            'tipo': tipo,
            'status': status,
        },
        'status_choices': SolicitacaoDesignacao.STATUS_CHOICES,
    }
    
    return render(request, 'pages/minhas_solicitacoes.html', context)


# ============================================================
# 📥 EXPORTAÇÃO
# ============================================================
@login_required
def exportar_excel(request, tipo):
    """Exporta dados para Excel."""
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse as HR
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='8B0000')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def style_header(sheet, num_cols):
        for col in range(1, num_cols + 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
    
    if tipo == 'oficiais':
        ws.title = 'Oficiais'
        ws.append(['CPF', 'RG', 'Nome', 'Nome de Guerra', 'Posto', 'Quadro', 'OBM', 'Função', 'Email', 'Telefone', 'Score'])
        style_header(ws, 11)
        for o in Oficial.objects.all():
            ws.append([o.cpf, o.rg, o.nome, o.nome_guerra, o.posto, o.quadro, o.obm, o.funcao, o.email, o.telefone, o.score])
    
    elif tipo == 'missoes':
        ws.title = 'Missões'
        ws.append(['ID', 'Tipo', 'Nome', 'Descrição', 'Local', 'Data Início', 'Data Fim', 'Status', 'Documento'])
        style_header(ws, 9)
        for m in Missao.objects.all():
            ws.append([m.id, m.tipo, m.nome, m.descricao, m.local, 
                      m.data_inicio.strftime('%Y-%m-%d') if m.data_inicio else '', 
                      m.data_fim.strftime('%Y-%m-%d') if m.data_fim else '', 
                      m.status, m.documento_referencia])
    
    elif tipo == 'designacoes':
        ws.title = 'Designações'
        ws.append(['ID', 'ID Missão', 'Nome Missão', 'RG Oficial', 'Nome Oficial', 'Função', 'Complexidade', 'Observações'])
        style_header(ws, 8)
        
        # Verificar se está consultando outro oficial
        oficial_id = request.GET.get('oficial_id')
        
        if oficial_id:
            # Consultando outro oficial
            try:
                oficial_consulta = Oficial.objects.get(pk=oficial_id)
                if request.user.pode_ver_oficial(oficial_consulta):
                    designacoes = Designacao.objects.select_related('missao', 'oficial').filter(
                        oficial=oficial_consulta
                    )
                else:
                    designacoes = Designacao.objects.none()
            except Oficial.DoesNotExist:
                designacoes = Designacao.objects.none()
        elif request.user.oficial and not request.user.is_admin:
            # Próprio oficial (não admin)
            designacoes = Designacao.objects.select_related('missao', 'oficial').filter(
                oficial=request.user.oficial
            )
        else:
            # Admin vê todos
            designacoes = Designacao.objects.select_related('missao', 'oficial').all()
        
        for d in designacoes:
            ws.append([d.id, d.missao.id, d.missao.nome, d.oficial.rg, str(d.oficial), 
                      d.funcao_na_missao, d.complexidade, d.observacoes])
    
    elif tipo == 'unidades':
        ws.title = 'Unidades'
        ws.append(['ID', 'Nome', 'Sigla', 'Tipo', 'ID Comando Superior'])
        style_header(ws, 5)
        for u in Unidade.objects.all():
            ws.append([u.id, u.nome, u.sigla, u.tipo, u.comando_superior_id or ''])
    
    elif tipo == 'usuarios':
        ws.title = 'Usuários'
        ws.append(['ID', 'CPF', 'Perfil', 'RG Oficial', 'Nome Oficial', 'Ativo'])
        style_header(ws, 6)
        for u in Usuario.objects.select_related('oficial').all():
            ws.append([u.id, u.cpf, u.role, 
                      u.oficial.rg if u.oficial else '', 
                      str(u.oficial) if u.oficial else '', 
                      'Sim' if u.is_active else 'Não'])
    
    elif tipo == 'modelo':
        # Criar planilha modelo com todas as abas
        return gerar_modelo_importacao()
    
    # Ajustar largura das colunas
    for column_cells in ws.columns:
        length = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)
    
    response = HR(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=sigem_{tipo}.xlsx'
    wb.save(response)
    
    return response


def gerar_modelo_importacao():
    """Gera planilha modelo para importação."""
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse as HR
    
    wb = openpyxl.Workbook()
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='8B0000')
    info_font = Font(bold=True, color='8B0000')
    
    def setup_sheet(sheet, headers, widths, example, info_col, info_data):
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            sheet.column_dimensions[get_column_letter(col)].width = widths[col-1]
        
        for col, value in enumerate(example, 1):
            sheet.cell(row=2, column=col, value=value)
        
        if info_data:
            for row, (title, values) in enumerate(info_data.items(), 1):
                sheet.cell(row=row, column=info_col, value=title).font = info_font
                for i, v in enumerate(values, 1):
                    sheet.cell(row=row+i, column=info_col, value=v)
        
        sheet.freeze_panes = 'A2'
    
    # Aba Oficiais
    ws = wb.active
    ws.title = 'Oficiais'
    setup_sheet(ws,
        ['CPF*', 'RG*', 'Nome Completo*', 'Nome de Guerra', 'Posto*', 'Quadro*', 'OBM', 'Função', 'Email', 'Telefone'],
        [15, 15, 35, 20, 12, 15, 25, 25, 30, 18],
        ['12345678901', 'RG123456', 'JOÃO DA SILVA', 'SILVA', 'Cap', 'QOC', '1º BBM', 'Cmt Cia', 'joao@email.com', '62999999999'],
        12,
        {'POSTOS:': ['Cel', 'TC', 'Maj', 'Cap', '1º Ten', '2º Ten', 'Asp'],
         'QUADROS:': ['QOC', 'QOA/Adm', 'QOA/Mús', 'QOM/Médico', 'QOM/Dentista']}
    )
    
    # Aba Missões
    ws2 = wb.create_sheet('Missoes')
    setup_sheet(ws2,
        ['Tipo*', 'Nome*', 'Descrição', 'Local', 'Data Início', 'Data Fim', 'Status*', 'Documento'],
        [18, 35, 40, 25, 15, 15, 18, 20],
        ['OPERACIONAL', 'Operação Exemplo', 'Descrição da missão', 'Goiânia-GO', '2024-01-15', '2024-01-20', 'EM_ANDAMENTO', 'SEI-123'],
        10,
        {'TIPOS:': ['OPERACIONAL', 'ADMINISTRATIVA', 'ENSINO', 'CORREICIONAL', 'COMISSAO', 'ACAO_SOCIAL'],
         'STATUS:': ['PLANEJADA', 'EM_ANDAMENTO', 'CONCLUIDA', 'CANCELADA']}
    )
    
    # Aba Designações
    ws3 = wb.create_sheet('Designacoes')
    setup_sheet(ws3,
        ['ID Missão*', 'RG Oficial*', 'Função*', 'Complexidade*', 'Observações'],
        [15, 18, 20, 15, 40],
        [1, 'RG123456', 'COMANDANTE', 'ALTA', 'Observação opcional'],
        7,
        {'FUNÇÕES:': ['COMANDANTE', 'SUBCOMANDANTE', 'COORDENADOR', 'PRESIDENTE', 'MEMBRO', 'AUXILIAR', 'INSTRUTOR', 'ENCARREGADO', 'RELATOR', 'ESCRIVAO'],
         'COMPLEXIDADE:': ['BAIXA', 'MEDIA', 'ALTA']}
    )
    
    # Aba Unidades
    ws4 = wb.create_sheet('Unidades')
    setup_sheet(ws4,
        ['Nome*', 'Sigla', 'Tipo*', 'ID Cmd Superior'],
        [40, 15, 18, 18],
        ['1º Batalhão BM', '1º BBM', 'BBM', ''],
        6,
        {'TIPOS:': ['COMANDO_GERAL', 'DIRETORIA', 'BBM', 'CIBM', 'CBM', 'SECAO']}
    )
    
    # Aba Usuários
    ws5 = wb.create_sheet('Usuarios')
    setup_sheet(ws5,
        ['CPF*', 'Perfil*', 'RG Oficial Vinculado'],
        [18, 15, 22],
        ['12345678901', 'oficial', 'RG123456'],
        5,
        {'PERFIS:': ['admin', 'gestor', 'comandante', 'oficial'],
         'NOTA:': ['Senha padrão: 123456']}
    )
    
    response = HR(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=sigem_modelo_importacao.xlsx'
    wb.save(response)
    
    return response


@login_required
def exportar_pdf(request, tipo):
    """Exporta dados para PDF - Relatório de designações do oficial."""
    
    from django.http import HttpResponse
    from django.conf import settings
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from io import BytesIO
    from datetime import datetime
    import os
    
    # Só permite PDF de designações por enquanto
    if tipo != 'designacoes':
        messages.info(request, 'Exportação PDF disponível apenas para designações.')
        return redirect(request.META.get('HTTP_REFERER', 'dashboard'))
    
    # Verificar se está consultando outro oficial
    oficial_id = request.GET.get('oficial_id')
    
    if oficial_id:
        # Consultando outro oficial
        oficial = get_object_or_404(Oficial, pk=oficial_id)
        
        # Verificar permissão
        if not request.user.pode_ver_oficial(oficial):
            messages.error(request, 'Você não tem permissão para gerar relatório deste oficial.')
            return redirect('consultar_oficial')
    else:
        # Próprio oficial
        if not request.user.oficial:
            messages.error(request, 'Usuário não vinculado a um oficial.')
            return redirect('consultar_oficial')
        oficial = request.user.oficial
    
    # Buscar designações do oficial
    designacoes = Designacao.objects.select_related('missao').filter(
        oficial=oficial
    ).order_by('-missao__status', '-criado_em')
    
    # Criar PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=1.5*cm,
        bottomMargin=1.5*cm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # ============================================================
    # FUNÇÃO AUXILIAR: Redimensionar imagem mantendo proporção
    # ============================================================
    def get_image_with_aspect_ratio(img_path, max_width, max_height, preserve_transparency=False):
        """Retorna Image do ReportLab mantendo proporção."""
        from PIL import Image as PILImage
        from PIL import ExifTags
        
        pil_img = PILImage.open(img_path)
        
        # Corrigir orientação EXIF
        try:
            for orientation in ExifTags.TAGS.keys():
                if ExifTags.TAGS[orientation] == 'Orientation':
                    break
            
            exif = pil_img._getexif()
            if exif is not None:
                orientation_value = exif.get(orientation)
                
                if orientation_value == 2:
                    pil_img = pil_img.transpose(PILImage.FLIP_LEFT_RIGHT)
                elif orientation_value == 3:
                    pil_img = pil_img.rotate(180)
                elif orientation_value == 4:
                    pil_img = pil_img.rotate(180).transpose(PILImage.FLIP_LEFT_RIGHT)
                elif orientation_value == 5:
                    pil_img = pil_img.rotate(-90, expand=True).transpose(PILImage.FLIP_LEFT_RIGHT)
                elif orientation_value == 6:
                    pil_img = pil_img.rotate(-90, expand=True)
                elif orientation_value == 7:
                    pil_img = pil_img.rotate(90, expand=True).transpose(PILImage.FLIP_LEFT_RIGHT)
                elif orientation_value == 8:
                    pil_img = pil_img.rotate(90, expand=True)
        except (AttributeError, KeyError, IndexError):
            pass
        
        # Obter dimensões originais
        orig_width, orig_height = pil_img.size
        
        # Calcular proporção mantendo aspect ratio
        ratio = min(max_width / orig_width, max_height / orig_height)
        new_width = orig_width * ratio
        new_height = orig_height * ratio
        
        # Salvar em buffer
        img_buffer = BytesIO()
        if preserve_transparency and pil_img.mode in ('RGBA', 'P', 'LA'):
            # Manter transparência - salvar como PNG
            pil_img.save(img_buffer, format='PNG')
        else:
            # Sem transparência - salvar como JPEG
            if pil_img.mode in ('RGBA', 'P', 'LA'):
                pil_img = pil_img.convert('RGB')
            pil_img.save(img_buffer, format='JPEG', quality=85)
        img_buffer.seek(0)
        
        return Image(img_buffer, width=new_width, height=new_height)
    
    # Estilos customizados
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.HexColor('#8B0000'),
        alignment=TA_LEFT,
        spaceAfter=2,
        leading=16
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.gray,
        alignment=TA_LEFT,
        spaceAfter=0
    )
    
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=4
    )
    
    # ============================================================
    # CABEÇALHO COM LOGO
    # ============================================================
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_cbmgo.png')
    
    # Criar elementos do cabeçalho
    titulo_principal = Paragraph("CORPO DE BOMBEIROS MILITAR<br/>DO ESTADO DE GOIÁS", title_style)
    subtitulo = Paragraph("Sistema de Gestão de Missões - SIGEM", subtitle_style)
    
    # Verificar se a logo existe
    if os.path.exists(logo_path):
        try:
            # Logo com proporção mantida (max 2cm x 2cm) - preserva transparência
            logo = get_image_with_aspect_ratio(logo_path, 2*cm, 2*cm, preserve_transparency=True)
            # Tabela: Logo à esquerda, Título à direita
            header_data = [[logo, [titulo_principal, subtitulo]]]
            header_table = Table(header_data, colWidths=[2.5*cm, 14.5*cm])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('ALIGN', (1, 0), (1, 0), 'LEFT'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))
        except Exception:
            # Se der erro na logo, título centralizado
            title_style.alignment = TA_CENTER
            subtitle_style.alignment = TA_CENTER
            header_data = [[[titulo_principal, subtitulo]]]
            header_table = Table(header_data, colWidths=[17*cm])
            header_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ]))
    else:
        # Sem logo - apenas título centralizado
        title_style.alignment = TA_CENTER
        subtitle_style.alignment = TA_CENTER
        header_data = [[[titulo_principal, subtitulo]]]
        header_table = Table(header_data, colWidths=[17*cm])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))
    
    elements.append(header_table)
    elements.append(Spacer(1, 0.5*cm))
    
    # Linha separadora
    linha_sep = Table([['']], colWidths=[17*cm], rowHeights=[2])
    linha_sep.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#8B0000')),
    ]))
    elements.append(linha_sep)
    elements.append(Spacer(1, 0.5*cm))
    
    # ============================================================
    # TÍTULO DO RELATÓRIO
    # ============================================================
    elements.append(Paragraph(
        "<b>RELATÓRIO DE DESIGNAÇÕES</b>", 
        ParagraphStyle('ReportTitle', fontSize=12, alignment=TA_CENTER, spaceAfter=15, textColor=colors.HexColor('#8B0000'))
    ))
    
    # ============================================================
    # DADOS DO OFICIAL COM FOTO
    # ============================================================
    # Verificar se oficial tem foto
    foto_path = None
    if oficial.foto:
        foto_path = oficial.foto.path if hasattr(oficial.foto, 'path') else None
    
    # Se não tem foto, usar avatar padrão
    if not foto_path or not os.path.exists(foto_path):
        foto_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'default_avatar.png')
    
    # Informações do oficial
    info_oficial = []
    info_oficial.append(Paragraph(f"<b>{oficial.posto} {oficial.nome}</b>", 
                                   ParagraphStyle('OficialNome', fontSize=11, spaceAfter=4)))
    info_oficial.append(Paragraph(f"<b>RG:</b> {oficial.rg}", info_style))
    info_oficial.append(Paragraph(f"<b>Quadro:</b> {oficial.quadro}", info_style))
    info_oficial.append(Paragraph(f"<b>OBM:</b> {oficial.obm or 'Não informado'}", info_style))
    info_oficial.append(Paragraph(f"<b>Data do Relatório:</b> {datetime.now().strftime('%d/%m/%Y às %H:%M')}", info_style))
    
    # Criar tabela com foto e informações
    if os.path.exists(foto_path):
        try:
            # Foto com proporção mantida (max 2.5cm x 3cm)
            foto_oficial = get_image_with_aspect_ratio(foto_path, 2.5*cm, 3*cm, preserve_transparency=False)
            oficial_data = [[foto_oficial, info_oficial]]
            oficial_table = Table(oficial_data, colWidths=[3*cm, 14*cm])
            oficial_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))
            elements.append(oficial_table)
        except Exception as e:
            # Se der erro na foto, mostrar só as informações
            for info in info_oficial:
                elements.append(info)
    else:
        for info in info_oficial:
            elements.append(info)
    
    elements.append(Spacer(1, 0.5*cm))
    
    # Resumo
    total_ativas = designacoes.filter(missao__status='EM_ANDAMENTO').count()
    total_baixa = designacoes.filter(missao__status='EM_ANDAMENTO', complexidade='BAIXA').count()
    total_media = designacoes.filter(missao__status='EM_ANDAMENTO', complexidade='MEDIA').count()
    total_alta = designacoes.filter(missao__status='EM_ANDAMENTO', complexidade='ALTA').count()
    
    resumo_data = [
        ['RESUMO DE MISSÕES EM ANDAMENTO', '', '', ''],
        ['Total Ativas', 'Baixa Complexidade', 'Média Complexidade', 'Alta Complexidade'],
        [str(total_ativas), str(total_baixa), str(total_media), str(total_alta)],
    ]
    
    resumo_table = Table(resumo_data, colWidths=[4*cm, 4*cm, 4*cm, 4*cm])
    resumo_table.setStyle(TableStyle([
        ('SPAN', (0, 0), (-1, 0)),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B0000')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f3f4f6')),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(resumo_table)
    elements.append(Spacer(1, 0.7*cm))
    
    # Tabela de designações
    if designacoes.exists():
        elements.append(Paragraph("<b>DETALHAMENTO DAS DESIGNAÇÕES</b>", ParagraphStyle('Heading', fontSize=11, spaceAfter=10)))
        
        table_data = [['Missão', 'Função', 'Complexidade', 'Status', 'Período']]
        
        for d in designacoes:
            periodo = ''
            if d.missao.data_inicio:
                periodo = d.missao.data_inicio.strftime('%d/%m/%Y')
                if d.missao.data_fim:
                    periodo += f" - {d.missao.data_fim.strftime('%d/%m/%Y')}"
            
            table_data.append([
                d.missao.nome[:40] + '...' if len(d.missao.nome) > 40 else d.missao.nome,
                d.get_funcao_na_missao_display(),
                d.get_complexidade_display(),
                d.missao.get_status_display(),
                periodo or '-'
            ])
        
        designacoes_table = Table(table_data, colWidths=[6*cm, 3*cm, 2.5*cm, 2.5*cm, 3*cm])
        designacoes_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B0000')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]))
        elements.append(designacoes_table)
    else:
        elements.append(Paragraph("Nenhuma designação encontrada.", info_style))
    
    # Rodapé
    elements.append(Spacer(1, 1*cm))
    elements.append(Paragraph(
        f"Documento gerado pelo SIGEM em {datetime.now().strftime('%d/%m/%Y às %H:%M')}",
        ParagraphStyle('Footer', fontSize=8, textColor=colors.gray, alignment=TA_CENTER)
    ))
    
    # Gerar PDF
    doc.build(elements)
    
    # Retornar resposta
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=relatorio_designacoes_{oficial.rg}.pdf'
    
    return response


# ============================================================
# 📤 IMPORTAÇÃO EM MASSA
# ============================================================
@login_required
@require_POST
def importar_excel(request, tipo):
    """Importa dados de arquivo Excel."""
    
    if not request.user.is_admin:
        messages.error(request, 'Sem permissão.')
        return redirect('admin_painel')
    
    import openpyxl
    from datetime import datetime
    
    arquivo = request.FILES.get('arquivo')
    
    if not arquivo:
        messages.error(request, 'Nenhum arquivo enviado.')
        return redirect('admin_painel')
    
    try:
        wb = openpyxl.load_workbook(arquivo)
        ws = wb.active
        
        count = 0
        errors = []
        
        # ============================================================
        # IMPORTAR OFICIAIS
        # ============================================================
        if tipo == 'oficiais':
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0]:  # Se tem CPF
                    try:
                        cpf = str(row[0]).replace('.', '').replace('-', '').strip()
                        oficial, created = Oficial.objects.update_or_create(
                            cpf=cpf,
                            defaults={
                                'rg': str(row[1]).strip() if row[1] else '',
                                'nome': str(row[2]).strip() if row[2] else '',
                                'nome_guerra': str(row[3]).strip() if row[3] else '',
                                'posto': str(row[4]).strip() if row[4] else '',
                                'quadro': str(row[5]).strip() if row[5] else '',
                                'obm': str(row[6]).strip() if row[6] else '',
                                'funcao': str(row[7]).strip() if row[7] else '',
                                'email': str(row[8]).strip() if row[8] else '',
                                'telefone': str(row[9]).strip() if row[9] else '',
                            }
                        )
                        count += 1
                        
                        # Criar usuário automaticamente se não existir
                        if created and not Usuario.objects.filter(cpf=cpf).exists():
                            Usuario.objects.create_user(
                                cpf=cpf,
                                password='123456',
                                oficial=oficial,
                                role='oficial'
                            )
                    except Exception as e:
                        errors.append(f'Linha {row_num}: {str(e)}')
        
        # ============================================================
        # IMPORTAR MISSÕES
        # ============================================================
        elif tipo == 'missoes':
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] and row[1]:  # Se tem tipo e nome
                    try:
                        # Processar datas
                        data_inicio = None
                        data_fim = None
                        
                        if row[4]:
                            if isinstance(row[4], datetime):
                                data_inicio = row[4].date()
                            else:
                                data_inicio = datetime.strptime(str(row[4]), '%Y-%m-%d').date()
                        
                        if row[5]:
                            if isinstance(row[5], datetime):
                                data_fim = row[5].date()
                            else:
                                data_fim = datetime.strptime(str(row[5]), '%Y-%m-%d').date()
                        
                        Missao.objects.create(
                            tipo=str(row[0]).strip().upper(),
                            nome=str(row[1]).strip(),
                            descricao=str(row[2]).strip() if row[2] else '',
                            local=str(row[3]).strip() if row[3] else '',
                            data_inicio=data_inicio,
                            data_fim=data_fim,
                            status=str(row[6]).strip().upper() if row[6] else 'PLANEJADA',
                            documento_referencia=str(row[7]).strip() if row[7] else '',
                        )
                        count += 1
                    except Exception as e:
                        errors.append(f'Linha {row_num}: {str(e)}')
        
        # ============================================================
        # IMPORTAR DESIGNAÇÕES
        # ============================================================
        elif tipo == 'designacoes':
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] and row[1]:  # Se tem missao_id e oficial_rg
                    try:
                        missao_id = int(row[0])
                        oficial_rg = str(row[1]).strip()
                        
                        # Buscar missão e oficial
                        missao = Missao.objects.get(id=missao_id)
                        oficial = Oficial.objects.get(rg=oficial_rg)
                        
                        Designacao.objects.update_or_create(
                            missao=missao,
                            oficial=oficial,
                            defaults={
                                'funcao_na_missao': str(row[2]).strip().upper() if row[2] else 'MEMBRO',
                                'complexidade': str(row[3]).strip().upper() if row[3] else 'MEDIA',
                                'observacoes': str(row[4]).strip() if row[4] else '',
                            }
                        )
                        count += 1
                    except Missao.DoesNotExist:
                        errors.append(f'Linha {row_num}: Missão ID {row[0]} não encontrada')
                    except Oficial.DoesNotExist:
                        errors.append(f'Linha {row_num}: Oficial RG {row[1]} não encontrado')
                    except Exception as e:
                        errors.append(f'Linha {row_num}: {str(e)}')
        
        # ============================================================
        # IMPORTAR UNIDADES
        # ============================================================
        elif tipo == 'unidades':
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0]:  # Se tem nome
                    try:
                        comando_superior = None
                        if row[3]:
                            try:
                                comando_superior = Unidade.objects.get(id=int(row[3]))
                            except Unidade.DoesNotExist:
                                pass
                        
                        Unidade.objects.update_or_create(
                            nome=str(row[0]).strip(),
                            defaults={
                                'sigla': str(row[1]).strip() if row[1] else '',
                                'tipo': str(row[2]).strip().upper() if row[2] else '',
                                'comando_superior': comando_superior,
                            }
                        )
                        count += 1
                    except Exception as e:
                        errors.append(f'Linha {row_num}: {str(e)}')
        
        # ============================================================
        # IMPORTAR USUÁRIOS
        # ============================================================
        elif tipo == 'usuarios':
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0]:  # Se tem CPF
                    try:
                        cpf = str(row[0]).replace('.', '').replace('-', '').strip()
                        role = str(row[1]).strip().lower() if row[1] else 'oficial'
                        
                        # Buscar oficial vinculado pelo RG
                        oficial = None
                        if row[2]:
                            try:
                                oficial = Oficial.objects.get(rg=str(row[2]).strip())
                            except Oficial.DoesNotExist:
                                pass
                        
                        if not Usuario.objects.filter(cpf=cpf).exists():
                            Usuario.objects.create_user(
                                cpf=cpf,
                                password='123456',
                                role=role,
                                oficial=oficial,
                            )
                            count += 1
                        else:
                            # Atualizar usuário existente
                            usuario = Usuario.objects.get(cpf=cpf)
                            usuario.role = role
                            if oficial:
                                usuario.oficial = oficial
                            usuario.save()
                            count += 1
                    except Exception as e:
                        errors.append(f'Linha {row_num}: {str(e)}')
        
        # Mensagem de resultado
        if count > 0:
            messages.success(request, f'{count} registros importados/atualizados com sucesso!')
        
        if errors:
            error_msg = f'Erros encontrados ({len(errors)}): ' + '; '.join(errors[:5])
            if len(errors) > 5:
                error_msg += f' ... e mais {len(errors) - 5} erros.'
            messages.warning(request, error_msg)
        
    except Exception as e:
        messages.error(request, f'Erro na importação: {str(e)}')
    
    return redirect('admin_painel')
