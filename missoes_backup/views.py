"""
============================================================
🎯 SIGEM - Views (Lógica das Páginas)
Sistema de Gestão de Missões - CBMGO
============================================================
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, Q, Avg
from django.utils import timezone
from django.views.decorators.http import require_POST, require_GET
from django.core.paginator import Paginator

from .models import Oficial, Missao, Designacao, Unidade, Usuario, SolicitacaoDesignacao


# ============================================================
# 🔐 AUTENTICAÇÃO
# ============================================================
def login_view(request):
    """Página de login."""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        cpf = request.POST.get('cpf', '').replace('.', '').replace('-', '')
        senha = request.POST.get('senha', '')
        
        user = authenticate(request, cpf=cpf, password=senha)
        
        if user is not None:
            login(request, user)
            user.ultimo_acesso = timezone.now()
            user.save(update_fields=['ultimo_acesso'])
            messages.success(request, f'Bem-vindo, {user}!')
            return redirect('dashboard')
        else:
            messages.error(request, 'CPF ou senha incorretos.')
    
    return render(request, 'auth/login.html')


@login_required
def logout_view(request):
    """Logout do usuário."""
    logout(request)
    messages.info(request, 'Você saiu do sistema.')
    return redirect('login')


# ============================================================
# 📊 DASHBOARD - VISÃO GERAL
# ============================================================
@login_required
def dashboard(request):
    """Página principal - Visão Geral."""
    
    # Estatísticas gerais
    total_oficiais = Oficial.objects.filter(ativo=True).count()
    total_missoes_ativas = Missao.objects.filter(status='EM_ANDAMENTO').count()
    total_missoes = Missao.objects.count()
    
    # Missões por tipo (apenas em andamento)
    missoes_por_tipo = Missao.objects.filter(status='EM_ANDAMENTO').values('tipo').annotate(
        total=Count('id')
    ).order_by('-total')
    
    # Oficiais mais escalados (top 10)
    oficiais_mais_escalados = Oficial.objects.filter(ativo=True).annotate(
        total_designacoes=Count('designacoes', filter=Q(designacoes__missao__status='EM_ANDAMENTO'))
    ).order_by('-total_designacoes')[:10]
    
    # Missões recentes
    missoes_recentes = Missao.objects.order_by('-criado_em')[:5]
    
    context = {
        'total_oficiais': total_oficiais,
        'total_missoes_ativas': total_missoes_ativas,
        'total_missoes': total_missoes,
        'missoes_por_tipo': missoes_por_tipo,
        'oficiais_mais_escalados': oficiais_mais_escalados,
        'missoes_recentes': missoes_recentes,
    }
    
    return render(request, 'pages/dashboard.html', context)


# ============================================================
# ⚖️ COMPARAR OFICIAIS
# ============================================================
@login_required
def comparar_oficiais(request):
    """Página para comparar carga de trabalho entre oficiais."""
    
    # Filtros disponíveis
    postos = Oficial.POSTO_CHOICES
    quadros = Oficial.QUADRO_CHOICES
    obms = Oficial.objects.values_list('obm', flat=True).distinct().order_by('obm')
    
    context = {
        'postos': postos,
        'quadros': quadros,
        'obms': obms,
    }
    
    return render(request, 'pages/comparar_oficiais.html', context)


# ============================================================
# 🗂️ DASHBOARD DE MISSÕES
# ============================================================
@login_required
def missoes_dashboard(request):
    """Dashboard completo de missões com organograma."""
    
    # Totalizadores por tipo (em andamento)
    tipos = ['OPERACIONAL', 'ADMINISTRATIVA', 'ENSINO', 'CORREICIONAL', 'COMISSAO', 'ACAO_SOCIAL']
    
    totais_por_tipo = []
    for tipo in tipos:
        total = Missao.objects.filter(tipo=tipo, status='EM_ANDAMENTO').count()
        totais_por_tipo.append({
            'tipo': tipo,
            'tipo_display': dict(Missao.TIPO_CHOICES).get(tipo, tipo),
            'total': total
        })
    
    # Filtros disponíveis
    status_choices = Missao.STATUS_CHOICES
    
    context = {
        'totais_por_tipo': totais_por_tipo,
        'status_choices': status_choices,
        'tipo_choices': Missao.TIPO_CHOICES,
    }
    
    return render(request, 'pages/missoes.html', context)


# ============================================================
# 👤 PAINEL DO OFICIAL
# ============================================================
@login_required
def painel_oficial(request):
    """Painel pessoal do oficial logado."""
    
    usuario = request.user
    oficial = usuario.oficial
    
    if not oficial:
        messages.warning(request, 'Seu usuário não está vinculado a um oficial.')
        return redirect('dashboard')
    
    # Designações do oficial
    designacoes = Designacao.objects.filter(oficial=oficial).select_related('missao').order_by('-criado_em')
    
    # Filtros
    tipo = request.GET.get('tipo', '')
    status = request.GET.get('status', '')
    complexidade = request.GET.get('complexidade', '')
    
    if tipo:
        designacoes = designacoes.filter(missao__tipo=tipo)
    if status:
        designacoes = designacoes.filter(missao__status=status)
    if complexidade:
        designacoes = designacoes.filter(complexidade=complexidade)
    
    # Solicitações pendentes
    solicitacoes = SolicitacaoDesignacao.objects.filter(solicitante=oficial).order_by('-criado_em')[:5]
    
    context = {
        'oficial': oficial,
        'designacoes': designacoes,
        'solicitacoes': solicitacoes,
        'tipo_choices': Missao.TIPO_CHOICES,
        'status_choices': Missao.STATUS_CHOICES,
        'complexidade_choices': Designacao.COMPLEXIDADE_CHOICES,
    }
    
    return render(request, 'pages/painel_oficial.html', context)


# ============================================================
# 🔧 PAINEL ADMINISTRATIVO
# ============================================================
@login_required
def admin_painel(request):
    """Painel administrativo com CRUD de todas as entidades."""
    
    # Verificar permissão
    if not request.user.is_gestor:
        messages.error(request, 'Acesso restrito a administradores e gestores.')
        return redirect('dashboard')
    
    context = {
        'aba_ativa': request.GET.get('aba', 'oficiais'),
    }
    
    return render(request, 'pages/admin_painel.html', context)


# ============================================================
# 🔄 HTMX - OFICIAIS
# ============================================================
@login_required
def htmx_oficiais_lista(request):
    """Retorna a lista de oficiais para a tabela administrativa."""
    
    oficiais = Oficial.objects.all().order_by('posto', 'nome')
    
    # Filtros
    posto = request.GET.get('posto', '')
    quadro = request.GET.get('quadro', '')
    obm = request.GET.get('obm', '')
    busca = request.GET.get('busca', '')
    
    if posto:
        oficiais = oficiais.filter(posto=posto)
    if quadro:
        oficiais = oficiais.filter(quadro=quadro)
    if obm:
        oficiais = oficiais.filter(obm__icontains=obm)
    if busca:
        oficiais = oficiais.filter(
            Q(nome__icontains=busca) | 
            Q(nome_guerra__icontains=busca) |
            Q(cpf__icontains=busca)
        )
    
    return render(request, 'htmx/oficiais_tabela.html', {'oficiais': oficiais})


@login_required
def htmx_oficiais_cards(request):
    """Retorna cards de oficiais para comparação."""
    
    ids = request.GET.get('ids', '')
    
    if ids:
        ids_list = [int(id) for id in ids.split(',') if id.isdigit()]
        oficiais = Oficial.objects.filter(id__in=ids_list).annotate(
            total_missoes=Count('designacoes', filter=Q(designacoes__missao__status='EM_ANDAMENTO')),
            total_chefia=Count('designacoes', filter=Q(
                designacoes__funcao_na_missao__in=['COMANDANTE', 'COORDENADOR', 'PRESIDENTE', 'ENCARREGADO'],
                designacoes__missao__status='EM_ANDAMENTO'
            ))
        )
    else:
        oficiais = []
    
    return render(request, 'htmx/oficiais_cards.html', {'oficiais': oficiais})


@login_required
def htmx_oficial_card(request, pk):
    """Retorna o card de um oficial específico."""
    
    oficial = get_object_or_404(Oficial, pk=pk)
    
    # Adiciona contagens
    oficial.total_missoes = oficial.designacoes.filter(missao__status='EM_ANDAMENTO').count()
    oficial.total_chefia = oficial.designacoes.filter(
        funcao_na_missao__in=['COMANDANTE', 'COORDENADOR', 'PRESIDENTE', 'ENCARREGADO'],
        missao__status='EM_ANDAMENTO'
    ).count()
    
    # Últimas missões
    oficial.ultimas_missoes = oficial.designacoes.select_related('missao').order_by('-criado_em')[:5]
    
    return render(request, 'htmx/oficial_card.html', {'oficial': oficial})


@login_required
@require_POST
def htmx_oficial_criar(request):
    """Cria um novo oficial via HTMX."""
    
    if not request.user.is_gestor:
        return HttpResponse('Sem permissão', status=403)
    
    try:
        oficial = Oficial.objects.create(
            cpf=request.POST.get('cpf', '').replace('.', '').replace('-', ''),
            rg=request.POST.get('rg', ''),
            nome=request.POST.get('nome', ''),
            nome_guerra=request.POST.get('nome_guerra', ''),
            posto=request.POST.get('posto', ''),
            quadro=request.POST.get('quadro', ''),
            obm=request.POST.get('obm', ''),
            funcao=request.POST.get('funcao', ''),
            email=request.POST.get('email', ''),
            telefone=request.POST.get('telefone', ''),
        )
        
        # Criar usuário automaticamente
        Usuario.objects.create_user(
            cpf=oficial.cpf,
            password='123456',
            oficial=oficial,
            role='oficial'
        )
        
        messages.success(request, f'Oficial {oficial} criado com sucesso!')
        
    except Exception as e:
        messages.error(request, f'Erro ao criar oficial: {str(e)}')
    
    return htmx_oficiais_lista(request)


@login_required
@require_POST
def htmx_oficial_editar(request, pk):
    """Edita um oficial via HTMX."""
    
    if not request.user.is_gestor:
        return HttpResponse('Sem permissão', status=403)
    
    oficial = get_object_or_404(Oficial, pk=pk)
    
    try:
        oficial.nome = request.POST.get('nome', oficial.nome)
        oficial.nome_guerra = request.POST.get('nome_guerra', oficial.nome_guerra)
        oficial.posto = request.POST.get('posto', oficial.posto)
        oficial.quadro = request.POST.get('quadro', oficial.quadro)
        oficial.obm = request.POST.get('obm', oficial.obm)
        oficial.funcao = request.POST.get('funcao', oficial.funcao)
        oficial.email = request.POST.get('email', oficial.email)
        oficial.telefone = request.POST.get('telefone', oficial.telefone)
        oficial.save()
        
        messages.success(request, f'Oficial {oficial} atualizado!')
        
    except Exception as e:
        messages.error(request, f'Erro ao atualizar: {str(e)}')
    
    return htmx_oficiais_lista(request)


@login_required
@require_POST
def htmx_oficial_excluir(request, pk):
    """Exclui um oficial via HTMX."""
    
    if not request.user.is_gestor:
        return HttpResponse('Sem permissão', status=403)
    
    oficial = get_object_or_404(Oficial, pk=pk)
    nome = str(oficial)
    
    try:
        oficial.delete()
        messages.success(request, f'Oficial {nome} excluído!')
    except Exception as e:
        messages.error(request, f'Erro ao excluir: {str(e)}')
    
    return htmx_oficiais_lista(request)


# ============================================================
# 🔄 HTMX - MISSÕES
# ============================================================
@login_required
def htmx_missoes_lista(request):
    """Retorna a lista de missões filtrada."""
    
    missoes = Missao.objects.all().order_by('-data_inicio')
    
    # Filtros
    tipo = request.GET.get('tipo', '')
    status = request.GET.get('status', '')
    local = request.GET.get('local', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    
    if tipo:
        missoes = missoes.filter(tipo=tipo)
    if status:
        missoes = missoes.filter(status=status)
    if local:
        missoes = missoes.filter(local__icontains=local)
    if data_inicio:
        missoes = missoes.filter(data_inicio__gte=data_inicio)
    if data_fim:
        missoes = missoes.filter(data_fim__lte=data_fim)
    
    return render(request, 'htmx/missoes_lista.html', {'missoes': missoes})


@login_required
def htmx_missao_organograma(request, pk):
    """Retorna o organograma de uma missão."""
    
    missao = get_object_or_404(Missao, pk=pk)
    
    designacoes = missao.designacoes.select_related('oficial').all()
    
    # Separar por hierarquia
    superiores = designacoes.filter(
        funcao_na_missao__in=['COMANDANTE', 'PRESIDENTE', 'COORDENADOR', 'ENCARREGADO']
    )
    subordinados = designacoes.exclude(
        funcao_na_missao__in=['COMANDANTE', 'PRESIDENTE', 'COORDENADOR', 'ENCARREGADO']
    )
    
    context = {
        'missao': missao,
        'superiores': superiores,
        'subordinados': subordinados,
    }
    
    return render(request, 'htmx/missao_organograma.html', context)


@login_required
@require_POST
def htmx_missao_criar(request):
    """Cria uma nova missão via HTMX."""
    
    if not request.user.is_gestor:
        return HttpResponse('Sem permissão', status=403)
    
    try:
        Missao.objects.create(
            tipo=request.POST.get('tipo', ''),
            nome=request.POST.get('nome', ''),
            descricao=request.POST.get('descricao', ''),
            local=request.POST.get('local', ''),
            data_inicio=request.POST.get('data_inicio') or None,
            data_fim=request.POST.get('data_fim') or None,
            status=request.POST.get('status', 'PLANEJADA'),
            documento_referencia=request.POST.get('documento_referencia', ''),
        )
        messages.success(request, 'Missão criada com sucesso!')
        
    except Exception as e:
        messages.error(request, f'Erro ao criar missão: {str(e)}')
    
    return htmx_missoes_lista(request)


@login_required
@require_POST
def htmx_missao_editar(request, pk):
    """Edita uma missão via HTMX."""
    
    if not request.user.is_gestor:
        return HttpResponse('Sem permissão', status=403)
    
    missao = get_object_or_404(Missao, pk=pk)
    
    try:
        missao.tipo = request.POST.get('tipo', missao.tipo)
        missao.nome = request.POST.get('nome', missao.nome)
        missao.descricao = request.POST.get('descricao', missao.descricao)
        missao.local = request.POST.get('local', missao.local)
        missao.status = request.POST.get('status', missao.status)
        missao.documento_referencia = request.POST.get('documento_referencia', missao.documento_referencia)
        
        data_inicio = request.POST.get('data_inicio')
        data_fim = request.POST.get('data_fim')
        if data_inicio:
            missao.data_inicio = data_inicio
        if data_fim:
            missao.data_fim = data_fim
        
        missao.save()
        messages.success(request, 'Missão atualizada!')
        
    except Exception as e:
        messages.error(request, f'Erro ao atualizar: {str(e)}')
    
    return htmx_missoes_lista(request)


@login_required
@require_POST
def htmx_missao_excluir(request, pk):
    """Exclui uma missão via HTMX."""
    
    if not request.user.is_gestor:
        return HttpResponse('Sem permissão', status=403)
    
    missao = get_object_or_404(Missao, pk=pk)
    nome = missao.nome
    
    try:
        missao.delete()
        messages.success(request, f'Missão "{nome}" excluída!')
    except Exception as e:
        messages.error(request, f'Erro ao excluir: {str(e)}')
    
    return htmx_missoes_lista(request)


# ============================================================
# 🔄 HTMX - DESIGNAÇÕES
# ============================================================
@login_required
def htmx_designacoes_lista(request):
    """Retorna a lista de designações."""
    
    designacoes = Designacao.objects.select_related('missao', 'oficial').all().order_by('-criado_em')
    
    return render(request, 'htmx/designacoes_tabela.html', {'designacoes': designacoes})


@login_required
@require_POST
def htmx_designacao_criar(request):
    """Cria uma nova designação via HTMX."""
    
    if not request.user.is_gestor:
        return HttpResponse('Sem permissão', status=403)
    
    try:
        missao_id = request.POST.get('missao_id')
        oficial_id = request.POST.get('oficial_id')
        
        Designacao.objects.create(
            missao_id=missao_id,
            oficial_id=oficial_id,
            funcao_na_missao=request.POST.get('funcao_na_missao', 'MEMBRO'),
            complexidade=request.POST.get('complexidade', 'MEDIA'),
            observacoes=request.POST.get('observacoes', ''),
        )
        messages.success(request, 'Designação criada!')
        
    except Exception as e:
        messages.error(request, f'Erro ao criar designação: {str(e)}')
    
    return htmx_designacoes_lista(request)


@login_required
@require_POST
def htmx_designacao_editar(request, pk):
    """Edita uma designação via HTMX."""
    
    if not request.user.is_gestor:
        return HttpResponse('Sem permissão', status=403)
    
    designacao = get_object_or_404(Designacao, pk=pk)
    
    try:
        designacao.funcao_na_missao = request.POST.get('funcao_na_missao', designacao.funcao_na_missao)
        designacao.complexidade = request.POST.get('complexidade', designacao.complexidade)
        designacao.observacoes = request.POST.get('observacoes', designacao.observacoes)
        designacao.save()
        messages.success(request, 'Designação atualizada!')
        
    except Exception as e:
        messages.error(request, f'Erro ao atualizar: {str(e)}')
    
    return htmx_designacoes_lista(request)


@login_required
@require_POST
def htmx_designacao_excluir(request, pk):
    """Exclui uma designação via HTMX."""
    
    if not request.user.is_gestor:
        return HttpResponse('Sem permissão', status=403)
    
    designacao = get_object_or_404(Designacao, pk=pk)
    
    try:
        designacao.delete()
        messages.success(request, 'Designação excluída!')
    except Exception as e:
        messages.error(request, f'Erro ao excluir: {str(e)}')
    
    return htmx_designacoes_lista(request)


# ============================================================
# 🔄 HTMX - UNIDADES
# ============================================================
@login_required
def htmx_unidades_lista(request):
    """Retorna a lista de unidades."""
    
    unidades = Unidade.objects.all().order_by('nome')
    
    return render(request, 'htmx/unidades_tabela.html', {'unidades': unidades})


@login_required
@require_POST
def htmx_unidade_criar(request):
    """Cria uma nova unidade via HTMX."""
    
    if not request.user.is_gestor:
        return HttpResponse('Sem permissão', status=403)
    
    try:
        comando_superior_id = request.POST.get('comando_superior_id')
        
        Unidade.objects.create(
            nome=request.POST.get('nome', ''),
            sigla=request.POST.get('sigla', ''),
            tipo=request.POST.get('tipo', ''),
            comando_superior_id=comando_superior_id if comando_superior_id else None,
        )
        messages.success(request, 'Unidade criada!')
        
    except Exception as e:
        messages.error(request, f'Erro ao criar unidade: {str(e)}')
    
    return htmx_unidades_lista(request)


@login_required
@require_POST
def htmx_unidade_editar(request, pk):
    """Edita uma unidade via HTMX."""
    
    if not request.user.is_gestor:
        return HttpResponse('Sem permissão', status=403)
    
    unidade = get_object_or_404(Unidade, pk=pk)
    
    try:
        unidade.nome = request.POST.get('nome', unidade.nome)
        unidade.sigla = request.POST.get('sigla', unidade.sigla)
        unidade.tipo = request.POST.get('tipo', unidade.tipo)
        
        comando_superior_id = request.POST.get('comando_superior_id')
        unidade.comando_superior_id = comando_superior_id if comando_superior_id else None
        
        unidade.save()
        messages.success(request, 'Unidade atualizada!')
        
    except Exception as e:
        messages.error(request, f'Erro ao atualizar: {str(e)}')
    
    return htmx_unidades_lista(request)


@login_required
@require_POST
def htmx_unidade_excluir(request, pk):
    """Exclui uma unidade via HTMX."""
    
    if not request.user.is_gestor:
        return HttpResponse('Sem permissão', status=403)
    
    unidade = get_object_or_404(Unidade, pk=pk)
    nome = unidade.nome
    
    try:
        unidade.delete()
        messages.success(request, f'Unidade "{nome}" excluída!')
    except Exception as e:
        messages.error(request, f'Erro ao excluir: {str(e)}')
    
    return htmx_unidades_lista(request)


# ============================================================
# 🔄 HTMX - USUÁRIOS
# ============================================================
@login_required
def htmx_usuarios_lista(request):
    """Retorna a lista de usuários."""
    
    if not request.user.is_admin:
        return HttpResponse('Sem permissão', status=403)
    
    usuarios = Usuario.objects.select_related('oficial').all().order_by('cpf')
    
    return render(request, 'htmx/usuarios_tabela.html', {'usuarios': usuarios})


@login_required
@require_POST
def htmx_usuario_criar(request):
    """Cria um novo usuário via HTMX."""
    
    if not request.user.is_admin:
        return HttpResponse('Sem permissão', status=403)
    
    try:
        cpf = request.POST.get('cpf', '').replace('.', '').replace('-', '')
        oficial_id = request.POST.get('oficial_id')
        
        Usuario.objects.create_user(
            cpf=cpf,
            password='123456',
            role=request.POST.get('role', 'oficial'),
            oficial_id=oficial_id if oficial_id else None,
        )
        messages.success(request, f'Usuário {cpf} criado com senha padrão 123456!')
        
    except Exception as e:
        messages.error(request, f'Erro ao criar usuário: {str(e)}')
    
    return htmx_usuarios_lista(request)


@login_required
@require_POST
def htmx_usuario_editar(request, pk):
    """Edita um usuário via HTMX."""
    
    if not request.user.is_admin:
        return HttpResponse('Sem permissão', status=403)
    
    usuario = get_object_or_404(Usuario, pk=pk)
    
    try:
        usuario.role = request.POST.get('role', usuario.role)
        usuario.is_active = request.POST.get('is_active') == 'on'
        usuario.save()
        messages.success(request, 'Usuário atualizado!')
        
    except Exception as e:
        messages.error(request, f'Erro ao atualizar: {str(e)}')
    
    return htmx_usuarios_lista(request)


@login_required
@require_POST
def htmx_usuario_excluir(request, pk):
    """Exclui um usuário via HTMX."""
    
    if not request.user.is_admin:
        return HttpResponse('Sem permissão', status=403)
    
    usuario = get_object_or_404(Usuario, pk=pk)
    
    if usuario == request.user:
        messages.error(request, 'Você não pode excluir seu próprio usuário!')
        return htmx_usuarios_lista(request)
    
    try:
        usuario.delete()
        messages.success(request, 'Usuário excluído!')
    except Exception as e:
        messages.error(request, f'Erro ao excluir: {str(e)}')
    
    return htmx_usuarios_lista(request)


@login_required
@require_POST
def htmx_usuario_reset_senha(request, pk):
    """Reseta a senha de um usuário para 123456."""
    
    if not request.user.is_admin:
        return HttpResponse('Sem permissão', status=403)
    
    usuario = get_object_or_404(Usuario, pk=pk)
    
    try:
        usuario.set_password('123456')
        usuario.save()
        messages.success(request, f'Senha do usuário {usuario.cpf} redefinida para 123456!')
    except Exception as e:
        messages.error(request, f'Erro ao resetar senha: {str(e)}')
    
    return htmx_usuarios_lista(request)


# ============================================================
# 🔄 HTMX - SOLICITAÇÕES
# ============================================================
@login_required
@require_POST
def htmx_solicitacao_criar(request):
    """Cria uma solicitação de designação."""
    
    if not request.user.oficial:
        messages.error(request, 'Usuário não vinculado a um oficial.')
        return HttpResponse(status=400)
    
    try:
        SolicitacaoDesignacao.objects.create(
            solicitante=request.user.oficial,
            nome_missao=request.POST.get('nome_missao', ''),
            funcao_na_missao=request.POST.get('funcao_na_missao', ''),
            complexidade=request.POST.get('complexidade', ''),
            documento_referencia=request.POST.get('documento_referencia', ''),
            justificativa=request.POST.get('justificativa', ''),
        )
        messages.success(request, 'Solicitação enviada com sucesso!')
        
    except Exception as e:
        messages.error(request, f'Erro ao criar solicitação: {str(e)}')
    
    return render(request, 'htmx/solicitacao_sucesso.html')


@login_required
def htmx_solicitacoes_lista(request):
    """Lista solicitações pendentes (para admin)."""
    
    if not request.user.is_gestor:
        return HttpResponse('Sem permissão', status=403)
    
    solicitacoes = SolicitacaoDesignacao.objects.filter(status='PENDENTE').select_related('solicitante').order_by('-criado_em')
    
    return render(request, 'htmx/solicitacoes_lista.html', {'solicitacoes': solicitacoes})


@login_required
@require_POST
def htmx_solicitacao_avaliar(request, pk):
    """Avalia uma solicitação (aprovar/recusar)."""
    
    if not request.user.is_gestor:
        return HttpResponse('Sem permissão', status=403)
    
    solicitacao = get_object_or_404(SolicitacaoDesignacao, pk=pk)
    
    acao = request.POST.get('acao')  # 'aprovar' ou 'recusar'
    
    try:
        solicitacao.status = 'APROVADA' if acao == 'aprovar' else 'RECUSADA'
        solicitacao.avaliado_por = request.user
        solicitacao.data_avaliacao = timezone.now()
        solicitacao.observacao_avaliador = request.POST.get('observacao', '')
        solicitacao.save()
        
        messages.success(request, f'Solicitação {solicitacao.get_status_display().lower()}!')
        
    except Exception as e:
        messages.error(request, f'Erro ao avaliar: {str(e)}')
    
    return htmx_solicitacoes_lista(request)


# ============================================================
# 📥 EXPORTAÇÃO
# ============================================================
@login_required
def exportar_excel(request, tipo):
    """Exporta dados para Excel."""
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse as HR
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='8B0000')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def style_header(sheet, num_cols):
        for col in range(1, num_cols + 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
    
    if tipo == 'oficiais':
        ws.title = 'Oficiais'
        ws.append(['CPF', 'RG', 'Nome', 'Nome de Guerra', 'Posto', 'Quadro', 'OBM', 'Função', 'Email', 'Telefone', 'Score'])
        style_header(ws, 11)
        for o in Oficial.objects.all():
            ws.append([o.cpf, o.rg, o.nome, o.nome_guerra, o.posto, o.quadro, o.obm, o.funcao, o.email, o.telefone, o.score])
    
    elif tipo == 'missoes':
        ws.title = 'Missões'
        ws.append(['ID', 'Tipo', 'Nome', 'Descrição', 'Local', 'Data Início', 'Data Fim', 'Status', 'Documento'])
        style_header(ws, 9)
        for m in Missao.objects.all():
            ws.append([m.id, m.tipo, m.nome, m.descricao, m.local, 
                      m.data_inicio.strftime('%Y-%m-%d') if m.data_inicio else '', 
                      m.data_fim.strftime('%Y-%m-%d') if m.data_fim else '', 
                      m.status, m.documento_referencia])
    
    elif tipo == 'designacoes':
        ws.title = 'Designações'
        ws.append(['ID', 'ID Missão', 'Nome Missão', 'RG Oficial', 'Nome Oficial', 'Função', 'Complexidade', 'Observações'])
        style_header(ws, 8)
        for d in Designacao.objects.select_related('missao', 'oficial').all():
            ws.append([d.id, d.missao.id, d.missao.nome, d.oficial.rg, str(d.oficial), 
                      d.funcao_na_missao, d.complexidade, d.observacoes])
    
    elif tipo == 'unidades':
        ws.title = 'Unidades'
        ws.append(['ID', 'Nome', 'Sigla', 'Tipo', 'ID Comando Superior'])
        style_header(ws, 5)
        for u in Unidade.objects.all():
            ws.append([u.id, u.nome, u.sigla, u.tipo, u.comando_superior_id or ''])
    
    elif tipo == 'usuarios':
        ws.title = 'Usuários'
        ws.append(['ID', 'CPF', 'Perfil', 'RG Oficial', 'Nome Oficial', 'Ativo'])
        style_header(ws, 6)
        for u in Usuario.objects.select_related('oficial').all():
            ws.append([u.id, u.cpf, u.role, 
                      u.oficial.rg if u.oficial else '', 
                      str(u.oficial) if u.oficial else '', 
                      'Sim' if u.is_active else 'Não'])
    
    elif tipo == 'modelo':
        # Criar planilha modelo com todas as abas
        return gerar_modelo_importacao()
    
    # Ajustar largura das colunas
    for column_cells in ws.columns:
        length = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)
    
    response = HR(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=sigem_{tipo}.xlsx'
    wb.save(response)
    
    return response


def gerar_modelo_importacao():
    """Gera planilha modelo para importação."""
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse as HR
    
    wb = openpyxl.Workbook()
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='8B0000')
    info_font = Font(bold=True, color='8B0000')
    
    def setup_sheet(sheet, headers, widths, example, info_col, info_data):
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            sheet.column_dimensions[get_column_letter(col)].width = widths[col-1]
        
        for col, value in enumerate(example, 1):
            sheet.cell(row=2, column=col, value=value)
        
        if info_data:
            for row, (title, values) in enumerate(info_data.items(), 1):
                sheet.cell(row=row, column=info_col, value=title).font = info_font
                for i, v in enumerate(values, 1):
                    sheet.cell(row=row+i, column=info_col, value=v)
        
        sheet.freeze_panes = 'A2'
    
    # Aba Oficiais
    ws = wb.active
    ws.title = 'Oficiais'
    setup_sheet(ws,
        ['CPF*', 'RG*', 'Nome Completo*', 'Nome de Guerra', 'Posto*', 'Quadro*', 'OBM', 'Função', 'Email', 'Telefone'],
        [15, 15, 35, 20, 12, 15, 25, 25, 30, 18],
        ['12345678901', 'RG123456', 'JOÃO DA SILVA', 'SILVA', 'Cap', 'QOC', '1º BBM', 'Cmt Cia', 'joao@email.com', '62999999999'],
        12,
        {'POSTOS:': ['Cel', 'Ten Cel', 'Maj', 'Cap', '1º Ten', '2º Ten', 'Asp'],
         'QUADROS:': ['QOC', 'QOA/Adm', 'QOA/Mús', 'QOM/Médico', 'QOM/Dentista']}
    )
    
    # Aba Missões
    ws2 = wb.create_sheet('Missoes')
    setup_sheet(ws2,
        ['Tipo*', 'Nome*', 'Descrição', 'Local', 'Data Início', 'Data Fim', 'Status*', 'Documento'],
        [18, 35, 40, 25, 15, 15, 18, 20],
        ['OPERACIONAL', 'Operação Exemplo', 'Descrição da missão', 'Goiânia-GO', '2024-01-15', '2024-01-20', 'EM_ANDAMENTO', 'SEI-123'],
        10,
        {'TIPOS:': ['OPERACIONAL', 'ADMINISTRATIVA', 'ENSINO', 'CORREICIONAL', 'COMISSAO', 'ACAO_SOCIAL'],
         'STATUS:': ['PLANEJADA', 'EM_ANDAMENTO', 'CONCLUIDA', 'CANCELADA']}
    )
    
    # Aba Designações
    ws3 = wb.create_sheet('Designacoes')
    setup_sheet(ws3,
        ['ID Missão*', 'RG Oficial*', 'Função*', 'Complexidade*', 'Observações'],
        [15, 18, 20, 15, 40],
        [1, 'RG123456', 'COMANDANTE', 'ALTA', 'Observação opcional'],
        7,
        {'FUNÇÕES:': ['COMANDANTE', 'SUBCOMANDANTE', 'COORDENADOR', 'PRESIDENTE', 'MEMBRO', 'AUXILIAR', 'INSTRUTOR', 'ENCARREGADO', 'RELATOR', 'ESCRIVAO'],
         'COMPLEXIDADE:': ['BAIXA', 'MEDIA', 'ALTA']}
    )
    
    # Aba Unidades
    ws4 = wb.create_sheet('Unidades')
    setup_sheet(ws4,
        ['Nome*', 'Sigla', 'Tipo*', 'ID Cmd Superior'],
        [40, 15, 18, 18],
        ['1º Batalhão BM', '1º BBM', 'BBM', ''],
        6,
        {'TIPOS:': ['COMANDO_GERAL', 'DIRETORIA', 'BBM', 'CIBM', 'CBM', 'SECAO']}
    )
    
    # Aba Usuários
    ws5 = wb.create_sheet('Usuarios')
    setup_sheet(ws5,
        ['CPF*', 'Perfil*', 'RG Oficial Vinculado'],
        [18, 15, 22],
        ['12345678901', 'oficial', 'RG123456'],
        5,
        {'PERFIS:': ['admin', 'gestor', 'comandante', 'oficial'],
         'NOTA:': ['Senha padrão: 123456']}
    )
    
    response = HR(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=sigem_modelo_importacao.xlsx'
    wb.save(response)
    
    return response


@login_required
def exportar_pdf(request, tipo):
    """Exporta dados para PDF."""
    # Implementação futura com reportlab ou weasyprint
    messages.info(request, 'Exportação PDF em desenvolvimento.')
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))


# ============================================================
# 📤 IMPORTAÇÃO EM MASSA
# ============================================================
@login_required
@require_POST
def importar_excel(request, tipo):
    """Importa dados de arquivo Excel."""
    
    if not request.user.is_admin:
        messages.error(request, 'Sem permissão.')
        return redirect('admin_painel')
    
    import openpyxl
    from datetime import datetime
    
    arquivo = request.FILES.get('arquivo')
    
    if not arquivo:
        messages.error(request, 'Nenhum arquivo enviado.')
        return redirect('admin_painel')
    
    try:
        wb = openpyxl.load_workbook(arquivo)
        ws = wb.active
        
        count = 0
        errors = []
        
        # ============================================================
        # IMPORTAR OFICIAIS
        # ============================================================
        if tipo == 'oficiais':
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0]:  # Se tem CPF
                    try:
                        cpf = str(row[0]).replace('.', '').replace('-', '').strip()
                        oficial, created = Oficial.objects.update_or_create(
                            cpf=cpf,
                            defaults={
                                'rg': str(row[1]).strip() if row[1] else '',
                                'nome': str(row[2]).strip() if row[2] else '',
                                'nome_guerra': str(row[3]).strip() if row[3] else '',
                                'posto': str(row[4]).strip() if row[4] else '',
                                'quadro': str(row[5]).strip() if row[5] else '',
                                'obm': str(row[6]).strip() if row[6] else '',
                                'funcao': str(row[7]).strip() if row[7] else '',
                                'email': str(row[8]).strip() if row[8] else '',
                                'telefone': str(row[9]).strip() if row[9] else '',
                            }
                        )
                        count += 1
                        
                        # Criar usuário automaticamente se não existir
                        if created and not Usuario.objects.filter(cpf=cpf).exists():
                            Usuario.objects.create_user(
                                cpf=cpf,
                                password='123456',
                                oficial=oficial,
                                role='oficial'
                            )
                    except Exception as e:
                        errors.append(f'Linha {row_num}: {str(e)}')
        
        # ============================================================
        # IMPORTAR MISSÕES
        # ============================================================
        elif tipo == 'missoes':
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] and row[1]:  # Se tem tipo e nome
                    try:
                        # Processar datas
                        data_inicio = None
                        data_fim = None
                        
                        if row[4]:
                            if isinstance(row[4], datetime):
                                data_inicio = row[4].date()
                            else:
                                data_inicio = datetime.strptime(str(row[4]), '%Y-%m-%d').date()
                        
                        if row[5]:
                            if isinstance(row[5], datetime):
                                data_fim = row[5].date()
                            else:
                                data_fim = datetime.strptime(str(row[5]), '%Y-%m-%d').date()
                        
                        Missao.objects.create(
                            tipo=str(row[0]).strip().upper(),
                            nome=str(row[1]).strip(),
                            descricao=str(row[2]).strip() if row[2] else '',
                            local=str(row[3]).strip() if row[3] else '',
                            data_inicio=data_inicio,
                            data_fim=data_fim,
                            status=str(row[6]).strip().upper() if row[6] else 'PLANEJADA',
                            documento_referencia=str(row[7]).strip() if row[7] else '',
                        )
                        count += 1
                    except Exception as e:
                        errors.append(f'Linha {row_num}: {str(e)}')
        
        # ============================================================
        # IMPORTAR DESIGNAÇÕES
        # ============================================================
        elif tipo == 'designacoes':
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] and row[1]:  # Se tem missao_id e oficial_rg
                    try:
                        missao_id = int(row[0])
                        oficial_rg = str(row[1]).strip()
                        
                        # Buscar missão e oficial
                        missao = Missao.objects.get(id=missao_id)
                        oficial = Oficial.objects.get(rg=oficial_rg)
                        
                        Designacao.objects.update_or_create(
                            missao=missao,
                            oficial=oficial,
                            defaults={
                                'funcao_na_missao': str(row[2]).strip().upper() if row[2] else 'MEMBRO',
                                'complexidade': str(row[3]).strip().upper() if row[3] else 'MEDIA',
                                'observacoes': str(row[4]).strip() if row[4] else '',
                            }
                        )
                        count += 1
                    except Missao.DoesNotExist:
                        errors.append(f'Linha {row_num}: Missão ID {row[0]} não encontrada')
                    except Oficial.DoesNotExist:
                        errors.append(f'Linha {row_num}: Oficial RG {row[1]} não encontrado')
                    except Exception as e:
                        errors.append(f'Linha {row_num}: {str(e)}')
        
        # ============================================================
        # IMPORTAR UNIDADES
        # ============================================================
        elif tipo == 'unidades':
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0]:  # Se tem nome
                    try:
                        comando_superior = None
                        if row[3]:
                            try:
                                comando_superior = Unidade.objects.get(id=int(row[3]))
                            except Unidade.DoesNotExist:
                                pass
                        
                        Unidade.objects.update_or_create(
                            nome=str(row[0]).strip(),
                            defaults={
                                'sigla': str(row[1]).strip() if row[1] else '',
                                'tipo': str(row[2]).strip().upper() if row[2] else '',
                                'comando_superior': comando_superior,
                            }
                        )
                        count += 1
                    except Exception as e:
                        errors.append(f'Linha {row_num}: {str(e)}')
        
        # ============================================================
        # IMPORTAR USUÁRIOS
        # ============================================================
        elif tipo == 'usuarios':
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0]:  # Se tem CPF
                    try:
                        cpf = str(row[0]).replace('.', '').replace('-', '').strip()
                        role = str(row[1]).strip().lower() if row[1] else 'oficial'
                        
                        # Buscar oficial vinculado pelo RG
                        oficial = None
                        if row[2]:
                            try:
                                oficial = Oficial.objects.get(rg=str(row[2]).strip())
                            except Oficial.DoesNotExist:
                                pass
                        
                        if not Usuario.objects.filter(cpf=cpf).exists():
                            Usuario.objects.create_user(
                                cpf=cpf,
                                password='123456',
                                role=role,
                                oficial=oficial,
                            )
                            count += 1
                        else:
                            # Atualizar usuário existente
                            usuario = Usuario.objects.get(cpf=cpf)
                            usuario.role = role
                            if oficial:
                                usuario.oficial = oficial
                            usuario.save()
                            count += 1
                    except Exception as e:
                        errors.append(f'Linha {row_num}: {str(e)}')
        
        # Mensagem de resultado
        if count > 0:
            messages.success(request, f'{count} registros importados/atualizados com sucesso!')
        
        if errors:
            error_msg = f'Erros encontrados ({len(errors)}): ' + '; '.join(errors[:5])
            if len(errors) > 5:
                error_msg += f' ... e mais {len(errors) - 5} erros.'
            messages.warning(request, error_msg)
        
    except Exception as e:
        messages.error(request, f'Erro na importação: {str(e)}')
    
    return redirect('admin_painel')